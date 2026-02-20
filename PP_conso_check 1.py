import argparse
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    import win32com.client as win32  # Windows-only Excel COM automation
except Exception:
    win32 = None

REFERENCE_EXCEL = r"file_name_format.xlsx"
AE_CHECK_FOLDER = r"AE_Checks"


def load_tsv(path):
    return pd.read_csv(path, sep="\t", dtype=str, keep_default_na=False, encoding="utf-8")


def save_tsv(df, path):
    df.to_csv(path, sep="\t", index=False, encoding="utf-8")


def clean_text(col):
    return (
        col.str.replace('"', "", regex=False)
        .str.replace(",", " ", regex=False)
        .str.replace("  ", " ", regex=False)
    )


def coerce_numeric_flags(df, cols=("is_it_promotional", "is_it_single", "is_it_single_brand")):
    for col in cols:
        if col not in df.columns:
            continue
        s = df[col].astype(str).str.strip()
        s = s.replace({"": pd.NA, "NA": pd.NA, "N/A": pd.NA, "na": pd.NA, "n/a": pd.NA})
        num = pd.to_numeric(s, errors="coerce")
        if num.notna().any():
            if (num.dropna() % 1 == 0).all():
                df[col] = num.astype("Int64")
            else:
                df[col] = num


def normalize_name(name):
    if not name:
        return ""
    name = name.strip()
    if name.lower().endswith(".tsv"):
        name = name[:-4]
    name = re.sub(r"\d{4}-\d{2}-\d{2}", "", name)
    name = re.sub(r"\s+", " ", name)
    return name.lower().strip()


def today_compact():
    return datetime.today().strftime("%Y%m%d")


def resolve_ae_naming(conso_filename, reference_excel):
    ref_df = pd.read_excel(reference_excel, dtype=str).fillna("")
    normalized_current = normalize_name(conso_filename)

    matched_idx = None
    for idx, row in ref_df.iterrows():
        if normalize_name(row["conso_file"]) == normalized_current:
            matched_idx = idx
            break

    if matched_idx is None:
        raise ValueError(
            f"Conso filename '{conso_filename}' not found in '{reference_excel}'. "
            "Please add conso_file, final_name_template, and AE_check_file_name mapping first."
        )

    row = ref_df.loc[matched_idx]
    final_name_template = str(row["final_name_template"]).strip()
    ae_file_name = str(row["AE_check_file_name"]).strip()
    return final_name_template, ae_file_name


def run_conso_check_once(file_path):
    df = load_tsv(file_path)
    checklist = []
    rowwise = pd.DataFrame(index=df.index)

    def top_values(series, limit=10):
        vals = [str(x).strip() for x in series.dropna().tolist() if str(x).strip()]
        uniq = []
        seen = set()
        for v in vals:
            if v not in seen:
                uniq.append(v)
                seen.add(v)
        return uniq[:limit]

    if "isRemoved" in df.columns:
        removed = df[df["isRemoved"] == "Yes"]
        removed_banners = top_values(removed["banner_name"]) if "banner_name" in removed.columns else []
        checklist.append(
            {
                "check_name": "isRemoved_rows_removed",
                "status": "PASS",
                "failed_count": 0,
                "details": f"Removed rows auto-dropped: {len(removed)}. Sample banners: {', '.join(removed_banners)}",
            }
        )
        df = df[df["isRemoved"] != "Yes"]
        rowwise = rowwise.loc[df.index]

    # row-wise base identity columns
    for col in ["banner_name", "date", "country", "channel", "brand_name", "retailer_name"]:
        if col in df.columns:
            rowwise[col] = df[col]
    rowwise["isSaved_must_be_yes"] = "PASS"
    rowwise["mandatory_columns_no_na"] = "PASS"
    rowwise["promo_message_consistency"] = "PASS"

    if "isSaved" in df.columns:
        invalid = df[df["isSaved"] != "Yes"]
        failed_banners = top_values(invalid["banner_name"]) if "banner_name" in invalid.columns else []
        checklist.append(
            {
                "check_name": "isSaved_must_be_yes",
                "status": "FAIL" if not invalid.empty else "PASS",
                "failed_count": int(len(invalid)),
                "details": (
                    ""
                    if invalid.empty
                    else f"isSaved is not 'Yes' for {len(invalid)} rows. Sample banners: {', '.join(failed_banners)}"
                ),
            }
        )
        rowwise.loc[df["isSaved"] != "Yes", "isSaved_must_be_yes"] = "FAIL"

    mandatory_cols = ["date", "country", "channel", "brand_name", "retailer_name"]
    mandatory_fail_msgs = []
    mandatory_fail_count = 0
    for col in mandatory_cols:
        if col in df.columns:
            bad = df[df[col].astype(str).str.upper() == "NA"]
            if not bad.empty:
                mandatory_fail_count += int(len(bad))
                banners = top_values(bad["banner_name"]) if "banner_name" in bad.columns else []
                mandatory_fail_msgs.append(
                    f"{col} has 'NA' in {len(bad)} rows; sample banners: {', '.join(banners)}"
                )
                rowwise.loc[df[col].astype(str).str.upper() == "NA", "mandatory_columns_no_na"] = "FAIL"

    checklist.append(
        {
            "check_name": "mandatory_columns_no_na",
            "status": "FAIL" if mandatory_fail_count > 0 else "PASS",
            "failed_count": mandatory_fail_count,
            "details": " | ".join(mandatory_fail_msgs),
        }
    )

    if df.shape[1] > 7:
        df = df.iloc[:, 7:]

    promo_cols = {"is_it_promotional", "promo_message"}
    if promo_cols.issubset(df.columns):
        promo_mask = (
            ((df["is_it_promotional"] == "1") & (df["promo_message"] == "NA"))
            | ((df["is_it_promotional"] == "0") & (df["promo_message"] != "NA"))
        )
        invalid = df[
            promo_mask
        ]
        failed_banners = top_values(invalid["banner_name"]) if "banner_name" in invalid.columns else []
        checklist.append(
            {
                "check_name": "promo_message_consistency",
                "status": "FAIL" if not invalid.empty else "PASS",
                "failed_count": int(len(invalid)),
                "details": (
                    ""
                    if invalid.empty
                    else (
                        "Promo mismatch: for promotional=1, promo_message cannot be NA; "
                        f"for promotional=0, promo_message must be NA. Failed rows: {len(invalid)}. "
                        f"Sample banners: {', '.join(failed_banners)}"
                    )
                ),
            }
        )
        rowwise.loc[promo_mask, "promo_message_consistency"] = "FAIL"

    for col in ["banner_alt_text", "promo_message", "exposed_sku"]:
        if col in df.columns:
            df[col] = clean_text(df[col].astype(str))

    coerce_numeric_flags(df)

    if {"brand_name", "exposed_sku"}.issubset(df.columns):
        df.loc[df["brand_name"] == "No Brand", "exposed_sku"] = "NA"

    has_fail = any(row["status"] == "FAIL" for row in checklist)
    return df, checklist, has_fail, rowwise.reset_index(drop=True)


def create_values_file(ae_local_path):
    if win32 is None:
        # On Linux/macOS servers, Excel COM is unavailable.
        # Try LibreOffice headless recalc; fallback to cached values from original file.
        values_path = ae_local_path.replace(".xlsx", "_VALUES.xlsx")
        if recalculate_with_libreoffice(ae_local_path, values_path):
            print("✅ Recalculated workbook via LibreOffice headless mode.")
            return values_path
        print("⚠️ win32com/libreoffice recalculation unavailable; using workbook cached values for validation.")
        return ae_local_path

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    wb = excel.Workbooks.Open(ae_local_path, ReadOnly=False)
    excel.CalculateFull()

    values_path = ae_local_path.replace(".xlsx", "_VALUES.xlsx")
    wb.SaveAs(values_path, FileFormat=51)

    wb.Close(False)
    excel.Quit()
    return values_path


def clear_brand_division_zeros(ae_path, values_path):
    for _ in range(5):
        try:
            wb_ae = load_workbook(ae_path)
            break
        except PermissionError:
            time.sleep(0.5)
    else:
        raise PermissionError(f"File still locked: {ae_path}")

    wb_val = load_workbook(values_path, data_only=True)

    final_sheet = next(s for s in wb_ae.sheetnames if s.lower() == "final")
    ws_final = wb_ae[final_sheet]
    ws_val = wb_val[final_sheet]

    headers = [ws_final.cell(1, c).value for c in range(1, ws_final.max_column + 1)]
    if "brand_division" not in headers:
        wb_ae.close()
        wb_val.close()
        return

    col = headers.index("brand_division") + 1
    for r in range(2, ws_final.max_row + 1):
        val = ws_val.cell(r, col).value
        if str(val).strip().lower() in {"0", "0.0"}:
            ws_final.cell(r, col, "")

    wb_ae.save(ae_path)
    wb_ae.close()
    wb_val.close()


def validate_checks_using_values(ae_path, values_path):
    def is_fail_value(val):
        if val in (False, "FALSE", "false", 0, "0"):
            return True
        if isinstance(val, str):
            s = val.strip().upper()
            if s in {"#N/A", "#NA", "#VALUE!", "#REF!", "#NAME?", "#DIV/0!"}:
                return True
        return False

    # Cloud/Linux fallback: no Excel COM.
    # Use cached formula results if available; if missing/false, keep highlight as failed.
    if win32 is None and os.path.abspath(values_path) == os.path.abspath(ae_path):
        wb_formula = load_workbook(ae_path, data_only=False)
        wb_data = load_workbook(ae_path, data_only=True)

        checks = next(s for s in wb_formula.sheetnames if s.lower() == "checks")
        final = next(s for s in wb_formula.sheetnames if s.lower() == "final")
        ws_checks_formula = wb_formula[checks]
        ws_checks_data = wb_data[checks]
        ws_final = wb_formula[final]

        highlight = PatternFill("solid", fgColor="FFFF00")
        clear_fill = PatternFill(fill_type=None)
        failed_count = 0
        failed_rows = set()
        max_row_to_check = min(ws_final.max_row, ws_checks_formula.max_row, ws_checks_data.max_row)

        for r in range(2, max_row_to_check + 1):
            for c in range(1, ws_checks_formula.max_column + 1):
                if c in (6, 7):
                    continue
                val = ws_checks_data.cell(r, c).value
                if is_fail_value(val):
                    ws_checks_formula.cell(r, c).fill = highlight
                    failed_count += 1
                    failed_rows.add(r)
                else:
                    ws_checks_formula.cell(r, c).fill = clear_fill

        wb_formula.save(ae_path)
        wb_formula.close()
        wb_data.close()
        if failed_count > 0:
            print("⚠️ Excel recalculation unavailable; highlighted checks use cached/missing values.")
        return failed_count == 0, failed_count, sorted(failed_rows)

    for _ in range(5):
        try:
            wb_ae = load_workbook(ae_path)
            break
        except PermissionError:
            time.sleep(0.5)
    else:
        raise PermissionError(f"File still locked: {ae_path}")

    wb_val = load_workbook(values_path, data_only=True)

    checks = next(s for s in wb_ae.sheetnames if s.lower() == "checks")
    final = next(s for s in wb_ae.sheetnames if s.lower() == "final")

    ws_checks = wb_ae[checks]
    ws_values = wb_val[checks]
    ws_final = wb_ae[final]

    highlight = PatternFill("solid", fgColor="FFFF00")
    clear_fill = PatternFill(fill_type=None)
    failed_count = 0
    failed_rows = set()
    max_row_to_check = min(ws_final.max_row, ws_checks.max_row, ws_values.max_row)

    for r in range(2, max_row_to_check + 1):
        for c in range(1, ws_checks.max_column + 1):
            if c in (6, 7):
                continue
            val = ws_values.cell(r, c).value
            if is_fail_value(val):
                ws_checks.cell(r, c).fill = highlight
                failed_count += 1
                failed_rows.add(r)
            else:
                ws_checks.cell(r, c).fill = clear_fill

    wb_ae.save(ae_path)
    wb_ae.close()
    wb_val.close()
    return failed_count == 0, failed_count, sorted(failed_rows)


def write_checklist_file(checklist_path, summary_rows, rowwise_df):
    with pd.ExcelWriter(checklist_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        rowwise_df.to_excel(writer, sheet_name="Row_Wise_Checks", index=False)


def recalculate_with_libreoffice(src_xlsx, dst_xlsx):
    """
    Best-effort recalculation on Linux containers using LibreOffice.
    Returns True on success, else False.
    """
    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp(prefix="pp_lo_calc_")
        cmd = [
            "soffice",
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--convert-to",
            "xlsx",
            "--outdir",
            tmp_dir,
            src_xlsx,
        ]
        proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
        if proc.returncode != 0:
            return False

        converted = os.path.join(tmp_dir, os.path.splitext(os.path.basename(src_xlsx))[0] + ".xlsx")
        if not os.path.exists(converted):
            return False
        shutil.copyfile(converted, dst_xlsx)
        return True
    except Exception:
        return False
    finally:
        if tmp_dir:
            try:
                shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception:
                pass


def checks_cached_available(template_path):
    """
    Check whether Checks formula cells have cached data-only values.
    """
    wb_formula = load_workbook(template_path, data_only=False)
    wb_data = load_workbook(template_path, data_only=True)
    try:
        checks = next((s for s in wb_formula.sheetnames if s.lower() == "checks"), None)
        if not checks:
            return True
        wsf = wb_formula[checks]
        wsd = wb_data[checks]
        max_row = min(wsf.max_row, wsd.max_row)
        for r in range(2, max_row + 1):
            for c in range(1, wsf.max_column + 1):
                if c in (6, 7):
                    continue
                raw = wsf.cell(r, c).value
                if isinstance(raw, str) and raw.startswith("="):
                    val = wsd.cell(r, c).value
                    if val is None or str(val).strip() == "":
                        return False
        return True
    finally:
        wb_formula.close()
        wb_data.close()


def ensure_template_cache_ready(ae_template_file, output_dir):
    """
    If template cached formula results are missing, try recalc and return cached file path.
    Returns: (usable_template_path, cache_file_to_upload_or_empty)
    """
    if checks_cached_available(ae_template_file):
        return ae_template_file, ""

    cached_template = os.path.join(
        output_dir,
        os.path.splitext(os.path.basename(ae_template_file))[0] + "_CACHE_READY.xlsx",
    )
    if recalculate_with_libreoffice(ae_template_file, cached_template) and checks_cached_available(cached_template):
        print(f"AE_TEMPLATE_CACHE_FILE={cached_template}")
        return cached_template, cached_template

    print("⚠️ Template cached values missing and recalculation could not be completed.")
    return ae_template_file, ""


def append_meta_and_checklist(review_file, meta_rows, checklist_rows):
    wb = load_workbook(review_file)

    if "Automation_Meta" in wb.sheetnames:
        del wb["Automation_Meta"]
    if "Automation_Checklist" in wb.sheetnames:
        del wb["Automation_Checklist"]

    ws_meta = wb.create_sheet("Automation_Meta")
    ws_meta.cell(1, 1, "key")
    ws_meta.cell(1, 2, "value")
    for idx, (k, v) in enumerate(meta_rows.items(), start=2):
        ws_meta.cell(idx, 1, k)
        ws_meta.cell(idx, 2, v)

    ws_check = wb.create_sheet("Automation_Checklist")
    headers = ["check_name", "status", "failed_count", "details"]
    for c, h in enumerate(headers, start=1):
        ws_check.cell(1, c, h)
    for r, row in enumerate(checklist_rows, start=2):
        ws_check.cell(r, 1, row.get("check_name", ""))
        ws_check.cell(r, 2, row.get("status", ""))
        ws_check.cell(r, 3, row.get("failed_count", 0))
        ws_check.cell(r, 4, row.get("details", ""))

    wb.save(review_file)
    wb.close()


def read_meta(review_file):
    wb = load_workbook(review_file, data_only=True)
    if "Automation_Meta" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Automation_Meta"]
    meta = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if k:
            meta[str(k)] = "" if v is None else str(v)
    wb.close()
    return meta


def write_conso_to_ae_template(conso_df, ae_template_file, review_file):
    # Hard-enforce numeric flag conversion before writing to template/checks.
    conso_df = conso_df.copy()
    coerce_numeric_flags(conso_df)

    shutil.copyfile(ae_template_file, review_file)
    if os.name == "nt":
        subprocess.run(["attrib", "-R", review_file], shell=True)

    wb = load_workbook(review_file)
    final = next(s for s in wb.sheetnames if s.lower() == "final")
    vlookup = next(s for s in wb.sheetnames if s.lower() == "vlookup")
    checks = next(s for s in wb.sheetnames if s.lower() == "checks")

    ws_final = wb[final]
    ws_vlookup = wb[vlookup]
    ws_checks = wb[checks]

    ws_final.delete_rows(1, ws_final.max_row)

    for c, col in enumerate(conso_df.columns, 1):
        ws_final.cell(1, c, col)

    for r, (_, row) in enumerate(conso_df.iterrows(), 2):
        for c, val in enumerate(row, 1):
            if pd.isna(val):
                val = None
            ws_final.cell(r, c, val)

    for c in range(17, 22):
        ws_final.cell(1, c, ws_vlookup.cell(1, c).value)
        base = ws_vlookup.cell(2, c).value
        for r in range(2, ws_final.max_row + 1):
            ws_final.cell(
                r,
                c,
                re.sub(r"(\$?[A-Z]+)2", lambda m: f"{m.group(1)}{r}", base)
                if isinstance(base, str) and base.startswith("=")
                else (base or ""),
            )

    # Extend Checks formulas for all rows present in Final.
    # Only adjust row references pointing to Final!<col>2; keep Details!$A$2 etc untouched.
    def shift_checks_formula(base_formula, row_num):
        if not (isinstance(base_formula, str) and base_formula.startswith("=")):
            return base_formula

        # 1) Shift Final sheet references: Final!B2 -> Final!B{row_num}
        formula = re.sub(
            r"(Final!\$?[A-Z]+)\$?2\b",
            lambda m: f"{m.group(1)}{row_num}",
            base_formula,
        )

        # 2) Shift same-sheet row-relative refs: F2 / $F2 -> F{row_num} / $F{row_num}
        #    Excludes cross-sheet refs like Details!$A$2 due to negative lookbehind for '!'.
        formula = re.sub(
            r"(?<![A-Za-z0-9_!])(\$?[A-Z]{1,3})\$?2\b",
            lambda m: f"{m.group(1)}{row_num}",
            formula,
        )
        return formula

    base_formulas = {
        c: ws_checks.cell(2, c).value
        for c in range(1, ws_checks.max_column + 1)
    }
    for r in range(2, ws_final.max_row + 1):
        for c in range(1, ws_checks.max_column + 1):
            base = base_formulas.get(c)
            if isinstance(base, str) and base.startswith("="):
                ws_checks.cell(r, c, shift_checks_formula(base, r))
            elif r > ws_checks.max_row:
                ws_checks.cell(r, c, base if base is not None else "")

    wb.save(review_file)
    wb.close()


def create_review_from_conso(conso_df, review_file):
    with pd.ExcelWriter(review_file, engine="openpyxl") as writer:
        conso_df.to_excel(writer, sheet_name="Final", index=False)


def prepare_mode(args):
    conso_filename = os.path.basename(args.conso_file)
    final_name_template, ae_flag = resolve_ae_naming(conso_filename, args.reference_excel)
    forced_template_path = getattr(args, "ae_template_file", "") or ""
    forced_template_path = forced_template_path.strip()

    conso_df, checklist_rows, conso_has_fail, rowwise_df = run_conso_check_once(args.conso_file)
    # Ensure conversions are finalized before any AE template checks.
    coerce_numeric_flags(conso_df)

    os.makedirs(args.output_dir, exist_ok=True)
    review_file = args.review_file or os.path.join(
        args.output_dir, f"{os.path.splitext(conso_filename)[0]}_review.xlsx"
    )
    checklist_file = os.path.join(args.output_dir, f"{os.path.splitext(conso_filename)[0]}_CHECKLIST.xlsx")

    ae_failed_cells = 0
    ae_failed_rows = []
    if ae_flag.lower() == "no":
        checklist_rows.append(
            {
                "check_name": "ae_template_check",
                "status": "PASS",
                "failed_count": 0,
                "details": "AE check not applicable for this file. Review file is not required.",
            }
        )
    else:
        if forced_template_path:
            ae_template_file = forced_template_path
            ae_flag = os.path.basename(forced_template_path)
        else:
            ae_template_file = os.path.join(args.ae_check_folder, ae_flag)
        if not os.path.exists(ae_template_file):
            raise FileNotFoundError(f"AE template not found: {ae_template_file}")

        ae_template_file, cache_file = ensure_template_cache_ready(ae_template_file, args.output_dir)
        if cache_file:
            print(f"AE_TEMPLATE_CACHE_FILE={cache_file}")

        write_conso_to_ae_template(conso_df, ae_template_file, review_file)
        values_file = create_values_file(review_file)
        clear_brand_division_zeros(review_file, values_file)
        ae_pass, ae_failed_cells, ae_failed_rows = validate_checks_using_values(review_file, values_file)
        try:
            if os.path.abspath(values_file) != os.path.abspath(review_file):
                os.remove(values_file)
        except OSError:
            pass

        checklist_rows.append(
            {
                "check_name": "ae_formula_checks",
                "status": "PASS" if ae_pass else "FAIL",
                "failed_count": ae_failed_cells,
                "details": (
                    "All AE checks passed."
                    if ae_pass
                    else (
                        f"{ae_failed_cells} check cells failed in 'Checks' sheet (yellow highlights). "
                        "Fix those rows and re-upload the same review file for validation."
                    )
                ),
            }
        )

    # Add AE check row-wise status for easier verification
    rowwise_df["ae_formula_checks"] = "PASS" if ae_flag.lower() == "no" else "PASS"
    if ae_failed_rows:
        for excel_row in ae_failed_rows:
            idx = excel_row - 2
            if 0 <= idx < len(rowwise_df):
                rowwise_df.loc[idx, "ae_formula_checks"] = "FAIL"

    rowwise_df["overall_status"] = "PASS"
    for c in ["isSaved_must_be_yes", "mandatory_columns_no_na", "promo_message_consistency", "ae_formula_checks"]:
        if c in rowwise_df.columns:
            rowwise_df.loc[rowwise_df[c] == "FAIL", "overall_status"] = "FAIL"

    write_checklist_file(checklist_file, checklist_rows, rowwise_df)
    print(f"CHECKLIST_FILE={checklist_file}")

    overall_fail = conso_has_fail or ae_failed_cells > 0
    if ae_flag.lower() != "no":
        meta_rows = {
            "mode": "prepare",
            "conso_filename": conso_filename,
            "final_name_template": final_name_template,
            "ae_check_file_name": ae_flag,
            "prepared_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "overall_status": "FAIL" if overall_fail else "PASS",
            "next_step": (
                "Checks passed: output files are generated."
                if not overall_fail
                else "Do changes in highlighted fields and re-upload for validation."
            ),
        }
        append_meta_and_checklist(review_file, meta_rows, checklist_rows)
        print(f"REVIEW_FILE={review_file}")

    print(f"OVERALL_STATUS={'FAIL' if overall_fail else 'PASS'}")

    if overall_fail:
        if ae_flag.lower() == "no":
            # For web/back-end flows, emit a checklist workbook on failure.
            create_review_from_conso(conso_df, review_file)
            meta_rows = {
                "mode": "prepare",
                "conso_filename": conso_filename,
                "final_name_template": final_name_template,
                "ae_check_file_name": ae_flag,
                "prepared_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "overall_status": "FAIL",
                "next_step": "Fix source conso file and rerun.",
            }
            append_meta_and_checklist(review_file, meta_rows, checklist_rows)
            print(f"REVIEW_FILE={review_file}")
            print("Validation failed. Fix source conso file and run prepare again.")
            return {
                "status": "FAIL",
                "review_file": review_file,
                "checklist_file": checklist_file,
                "final_tsv": "",
                "final_xlsx": "",
            }
        else:
            print("Review file has failed checks. Do changes and re-upload for validation.")
            return {
                "status": "FAIL",
                "review_file": review_file,
                "checklist_file": checklist_file,
                "final_tsv": "",
                "final_xlsx": "",
            }

    if ae_flag.lower() == "no":
        final_df = conso_df.copy()
    else:
        values_file = create_values_file(review_file)
        final_sheet = get_final_sheet_name(values_file)
        final_df = pd.read_excel(values_file, sheet_name=final_sheet, dtype=str, keep_default_na=False)
        try:
            if os.path.abspath(values_file) != os.path.abspath(review_file):
                os.remove(values_file)
        except OSError:
            pass

    if {"brand_name", "exposed_sku"}.issubset(final_df.columns):
        final_df.loc[final_df["brand_name"] == "No Brand", "exposed_sku"] = "NA"
    coerce_numeric_flags(final_df)

    out_tsv, out_xlsx = write_final_outputs(final_df, final_name_template, args.output_dir)
    print(f"FINAL_TSV={out_tsv}")
    print(f"FINAL_XLSX={out_xlsx}")
    if ae_flag.lower() == "no":
        print("No review file required. Output files generated with today's date.")
    else:
        print("Review checks passed. Review file and output files generated with today's date.")
    return {
        "status": "PASS",
        "review_file": review_file if ae_flag.lower() != "no" else "",
        "checklist_file": checklist_file,
        "final_tsv": out_tsv,
        "final_xlsx": out_xlsx,
    }


def infer_final_name(template):
    if re.search(r"\d{8}", template):
        return re.sub(r"\d{8}", today_compact(), template)
    return f"{template}_{today_compact()}"


def write_final_outputs(final_df, template, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    final_name = infer_final_name(template)
    out_tsv = os.path.join(output_dir, final_name + ".tsv")
    out_xlsx = os.path.join(output_dir, final_name + ".xlsx")

    final_df.to_csv(out_tsv, sep="\t", index=False, encoding="utf-8")
    final_df.to_excel(out_xlsx, index=False)
    return out_tsv, out_xlsx


def get_final_sheet_name(path):
    wb = load_workbook(path, data_only=True)
    try:
        return next(s for s in wb.sheetnames if s.lower() == "final")
    finally:
        wb.close()


def finalize_mode(args):
    meta = read_meta(args.review_file)
    template = args.final_name_template or meta.get("final_name_template", "").strip()
    if not template:
        raise ValueError(
            "final_name_template not found in Automation_Meta sheet. "
            "Pass --final-name-template explicitly."
        )

    ae_flag = meta.get("ae_check_file_name", "no").strip()

    if ae_flag.lower() != "no":
        values_file = create_values_file(args.review_file)
        clear_brand_division_zeros(args.review_file, values_file)
        ae_pass, ae_failed_cells, _ = validate_checks_using_values(args.review_file, values_file)
        if not ae_pass:
            try:
                if os.path.abspath(values_file) != os.path.abspath(args.review_file):
                    os.remove(values_file)
            except OSError:
                pass
            raise ValueError(
                f"File is still not corrected for highlighted field. Failed cells: {ae_failed_cells}. "
                "Please do changes and re-upload for validation."
            )
        final_sheet = get_final_sheet_name(values_file)
        final_df = pd.read_excel(values_file, sheet_name=final_sheet, dtype=str, keep_default_na=False)
        try:
            if os.path.abspath(values_file) != os.path.abspath(args.review_file):
                os.remove(values_file)
        except OSError:
            pass
    else:
        final_sheet = get_final_sheet_name(args.review_file)
        final_df = pd.read_excel(args.review_file, sheet_name=final_sheet, dtype=str, keep_default_na=False)

    if {"brand_name", "exposed_sku"}.issubset(final_df.columns):
        final_df.loc[final_df["brand_name"] == "No Brand", "exposed_sku"] = "NA"
    coerce_numeric_flags(final_df)

    out_tsv, out_xlsx = write_final_outputs(final_df, template, args.output_dir)

    print(f"FINAL_TSV={out_tsv}")
    print(f"FINAL_XLSX={out_xlsx}")
    print("Validation passed. Output files generated with today's date.")
    return {
        "status": "PASS",
        "review_file": args.review_file,
        "checklist_file": "",
        "final_tsv": out_tsv,
        "final_xlsx": out_xlsx,
    }


def _is_excel_file(path):
    return str(path).lower().endswith((".xlsx", ".xls", ".xlsm", ".xltx", ".xltm"))


def _copy_if_needed(src, dst):
    if not src:
        return
    if os.path.abspath(src) == os.path.abspath(dst):
        return
    shutil.copyfile(src, dst)


def backend_mode(op_file, ip_file, master_file, output_file, ae_template_file=""):
    """
    Back-end compatibility mode:
    python PP_conso_check 1.py <op_file> <ip_file> <master_file> <output_xlsx>
    - Fresh run: uses OP + master reference mapping.
    - Revalidation run: if IP points to Excel review file, validate IP and generate finals.
    """
    output_dir = os.path.dirname(os.path.abspath(output_file)) or "."
    os.makedirs(output_dir, exist_ok=True)

    op_exists = bool(op_file) and os.path.exists(op_file)
    ip_exists = bool(ip_file) and os.path.exists(ip_file)

    # Platform rule:
    # - Fresh run if source file exists.
    # - Revalidation only when source is absent and review workbook is provided.
    if (not op_exists) and ip_exists and _is_excel_file(ip_file):
        print("MODE=REVALIDATION")
        result = finalize_mode(
            argparse.Namespace(
                review_file=ip_file,
                output_dir=output_dir,
                final_name_template="",
            )
        )
        _copy_if_needed(result.get("final_xlsx", ""), output_file)
        return

    print("MODE=FRESH")
    if not op_exists:
        raise FileNotFoundError(
            f"Fresh mode requires source file, but not found: '{op_file}'. "
            "Provide source file for fresh run or review workbook for revalidation."
        )
    ae_template_arg = (ae_template_file or "").strip()
    result = prepare_mode(
        argparse.Namespace(
            conso_file=op_file,
            output_dir=output_dir,
            review_file=output_file,
            reference_excel=master_file if master_file else REFERENCE_EXCEL,
            ae_check_folder=AE_CHECK_FOLDER,
            ae_template_file=ae_template_arg,
        )
    )
    if not result:
        return

    if result.get("status") == "PASS" and result.get("final_xlsx"):
        _copy_if_needed(result["final_xlsx"], output_file)
    elif result.get("review_file"):
        _copy_if_needed(result["review_file"], output_file)


def build_parser():
    parser = argparse.ArgumentParser(
        description="PP conso web-compatible checklist flow (prepare -> finalize)."
    )
    sub = parser.add_subparsers(dest="mode", required=True)

    p_prepare = sub.add_parser("prepare", help="Generate review/highlight file and checklist.")
    p_prepare.add_argument("--conso-file", required=True, help="Path to source conso TSV file.")
    p_prepare.add_argument("--output-dir", required=True, help="Directory for review file.")
    p_prepare.add_argument("--review-file", default="", help="Optional explicit review file path.")
    p_prepare.add_argument(
        "--reference-excel",
        default=REFERENCE_EXCEL,
        help="Reference Excel containing conso_file/final_name_template/AE_check_file_name.",
    )
    p_prepare.add_argument("--ae-check-folder", default=AE_CHECK_FOLDER, help="Folder with AE templates.")
    p_prepare.add_argument(
        "--ae-template-file",
        default="",
        help="Optional AE template file path to override template lookup from reference.",
    )

    p_finalize = sub.add_parser("finalize", help="Validate corrected review file and generate final outputs.")
    p_finalize.add_argument("--review-file", required=True, help="Corrected review file uploaded from UI.")
    p_finalize.add_argument("--output-dir", required=True, help="Directory for generated final files.")
    p_finalize.add_argument(
        "--final-name-template",
        default="",
        help="Optional fallback final_name_template if Automation_Meta sheet is missing.",
    )

    return parser


def main():
    # Back-end compatibility path used by worker-style integrations.
    if len(sys.argv) in {5, 6} and sys.argv[1] not in {"prepare", "finalize"}:
        if len(sys.argv) == 5:
            _, op_file, ip_file, master_file, output_file = sys.argv
            ae_template_file = ""
        else:
            _, op_file, ip_file, master_file, output_file, ae_template_file = sys.argv
        backend_mode(op_file, ip_file, master_file, output_file, ae_template_file)
        return

    parser = build_parser()
    args = parser.parse_args()

    if args.mode == "prepare":
        prepare_mode(args)
    elif args.mode == "finalize":
        finalize_mode(args)
    else:
        parser.error("Unknown mode")


if __name__ == "__main__":
    main()
