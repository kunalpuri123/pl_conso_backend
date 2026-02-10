import pandas as pd
from openpyxl import load_workbook
import os
import re
import sys
import zipfile
from datetime import datetime

print("=== INPUT CREATION STARTED ===", flush=True)

# =========================================================
# ARGUMENTS
# =========================================================

if len(sys.argv) != 5:
    print("Usage: python input_creation_script.py <BIZ> <CRAWL> <TEMPLATE> <OUTPUT>")
    sys.exit(1)

BIZ_FILE = sys.argv[1]
CRAWL_FILE = sys.argv[2]
TEMPLATE_FILE = sys.argv[3]
OUTPUT_FILE = sys.argv[4]

RUN_DIR = os.path.dirname(OUTPUT_FILE)
TSV_FOLDER = os.path.join(RUN_DIR, "tsv_files")

os.makedirs(TSV_FOLDER, exist_ok=True)

print("Biz:", BIZ_FILE)
print("Crawl:", CRAWL_FILE)
print("Template:", TEMPLATE_FILE)
print("Output:", OUTPUT_FILE)


# =========================================================
# HELPERS
# =========================================================

def sanitize(text):
    if text is None:
        return "NA"
    s = str(text)
    if s.strip().lower() in {"nan", "none", ""}:
        return "NA"
    return re.sub(r'[^\w\-_.]', '_', s)


# =========================================================
# 1️⃣ LOAD FILES
# =========================================================

print("Loading files...", flush=True)

biz_df = pd.read_excel(BIZ_FILE, engine="openpyxl")
crawl_df = pd.read_excel(CRAWL_FILE, engine="openpyxl")
template_df = pd.read_excel(TEMPLATE_FILE, engine="openpyxl")

biz_df.columns = biz_df.columns.str.strip()
crawl_df.columns = crawl_df.columns.str.strip()
template_df.columns = template_df.columns.str.strip()

biz_df = biz_df.fillna("n/a")

print(f"Biz rows: {len(biz_df)}")
print(f"Crawl rows: {len(crawl_df)}")


# =========================================================
# 2️⃣ MERGE
# =========================================================

merge_keys = ["Batch", "SubBatch", "scope_name"]

for key in merge_keys:
    if key not in biz_df.columns or key not in crawl_df.columns:
        raise Exception(f"Missing merge key: {key}")

merged_df = pd.merge(
    biz_df,
    crawl_df,
    on=merge_keys,
    how="outer",
    indicator=True
)

print("Merged rows:", len(merged_df))


# =========================================================
# 3️⃣ ALIGN WITH TEMPLATE
# =========================================================

final_df = pd.DataFrame(columns=template_df.columns)

# Case-insensitive column mapping to avoid empty columns due to casing differences
merged_col_map = {c.lower(): c for c in merged_df.columns}

for col in final_df.columns:
    src = merged_col_map.get(col.lower())
    if src:
        final_df[col] = merged_df[src]
    else:
        final_df[col] = None


# =========================================================
# 4️⃣ EXTRA COLUMNS
# =========================================================

final_df["Base_ID"] = range(1, len(final_df) + 1)
final_df["uniqueIdentifier"] = final_df["Base_ID"]

if "category_input" in final_df.columns:
    final_df["category_input"] = final_df["category_input"].fillna("n/a").replace("", "n/a")

for col in ["ValidateParsedOutput", "FetchNextCrawlURL"]:
    if col in final_df.columns:
        final_df[col] = final_df[col].astype(str).str.upper()

final_df["ip_search_keyword_category"] = "n/a"
final_df["ip_filter_by_category"] = "n/a"


# =========================================================
# 5️⃣ SAVE EXCEL (merged_output.xlsx)
# =========================================================

print("Saving merged Excel...", flush=True)

wb = load_workbook(TEMPLATE_FILE)
ws = wb.active

# clear old data
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.value = None

for r_idx, row in final_df.iterrows():
    for c_idx, value in enumerate(row):
        ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)

wb.save(OUTPUT_FILE)

print("Excel saved:", OUTPUT_FILE)


# =========================================================
# 6️⃣ CREATE TSVs (per batch)
# =========================================================

print("Generating TSVs...", flush=True)

biz_only_df = final_df[merged_df["_merge"].isin(["both", "left_only"])].copy()

tsv_paths = []

for (batch, subbatch), group_df in biz_only_df.groupby(["Batch", "SubBatch"]):

    scope = group_df.get("scope_name", ["NA"]).iloc[0]
    type_ = group_df.get("type", ["NA"]).iloc[0]
    channel = group_df.get("Channel", ["NA"]).iloc[0]

    filename = f"{sanitize(scope)}_{sanitize(subbatch)}_{sanitize(batch)}_{sanitize(type_)}_{sanitize(channel)}.tsv"

    filepath = os.path.join(TSV_FOLDER, filename)

    group_df.to_csv(filepath, sep="\t", index=False)

    tsv_paths.append(filepath)

    print("TSV created:", filename)


# =========================================================
# 7️⃣ CONSOLIDATED TSV
# =========================================================

today = datetime.today().strftime("%Y%m%d")
consolidated_name = f"LVMH_CP_CP_TBL_LVMH_PL_INPUT_Template_{today}.tsv"
consolidated_path = os.path.join(TSV_FOLDER, consolidated_name)

header_written = False

with open(consolidated_path, "w", encoding="utf-8") as outfile:
    for path in tsv_paths:
        with open(path, "r", encoding="utf-8") as infile:
            lines = infile.readlines()
            if not header_written:
                outfile.write(lines[0])
                header_written = True
            outfile.writelines(lines[1:])

tsv_paths.append(consolidated_path)

print("Consolidated TSV created")


# =========================================================
# 8️⃣ ZIP ALL TSVs
# =========================================================

zip_path = os.path.join(RUN_DIR, "all_tsv_files.zip")

print("Creating ZIP...", flush=True)

with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
    for path in tsv_paths:
        z.write(path, os.path.basename(path))

print("ZIP created:", zip_path)


print("=== INPUT CREATION COMPLETED SUCCESSFULLY ===", flush=True)
