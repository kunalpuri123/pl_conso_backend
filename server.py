from fastapi import FastAPI, BackgroundTasks, Depends, Header, HTTPException, Request, status
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware

from supabase import create_client
from dotenv import load_dotenv
from datetime import datetime
from collections import defaultdict, deque
from threading import Lock

import subprocess
import tempfile
import os
import shutil
import signal
import hashlib
import re

from pdf_report_generator import generate_pdf_from_ai_report
from ai_analyzer import analyze_output_with_gemini


# =========================================================
# ENV
# =========================================================

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise Exception("SUPABASE_URL or SUPABASE_SERVICE_KEY missing in .env")

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


# =========================================================
# APP
# =========================================================

ENABLE_API_DOCS = os.getenv("ENABLE_API_DOCS", "false").strip().lower() == "true"

app = FastAPI(
    docs_url="/docs" if ENABLE_API_DOCS else None,
    redoc_url=None,
    openapi_url="/openapi.json" if ENABLE_API_DOCS else None,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8080",
        "http://127.0.0.1:8080",
        "https://pl-conso-frontend.vercel.app"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


BASE_WORKDIR = "./work"
os.makedirs(BASE_WORKDIR, exist_ok=True)

RATE_LIMITS = {
    "trigger": (20, 60),       # 20 requests per 60s
    "logs": (60, 60),          # 60 requests per 60s
    "mutating_run": (10, 60),  # 10 requests per 60s
}
_rate_buckets = defaultdict(deque)
_rate_lock = Lock()


# =========================================================
# HELPERS
# =========================================================

def log(run_id: str, level: str, message: str):
    """
    Insert one log row into DB
    """
    if not message or message.strip() == "":
        return

    supabase.table("run_logs").insert({
        "run_id": run_id,
        "level": level,
        "message": message.strip()
    }).execute()


def download_from_storage(bucket, storage_path, local_path):
    def _download(path):
        data = supabase.storage.from_(bucket).download(path)
        if isinstance(data, dict) and data.get("statusCode"):
            raise Exception(data)
        if data is None or data == b"":
            raise Exception({"statusCode": 404, "error": "not_found", "message": "Empty response"})
        return data

    tried = []
    last_error = None

    for path in [storage_path, os.path.basename(storage_path)]:
        if not path or path in tried:
            continue
        tried.append(path)
        try:
            data = _download(path)
            with open(local_path, "wb") as f:
                f.write(data)
            return path
        except Exception as e:
            last_error = e

    raise Exception(f"Download failed for bucket '{bucket}'. Tried: {tried}. Last error: {last_error}")


def upload_to_storage(bucket, storage_path, local_path):
    with open(local_path, "rb") as f:
        res = supabase.storage.from_(bucket).upload(storage_path, f)

    print("UPLOAD RESULT:", res)   # 🔥 add this

def sha256_file(path, chunk_size=1024 * 1024):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(chunk_size), b""):
            h.update(chunk)
    return h.hexdigest()

def remove_from_buckets(storage_path, buckets):
    for b in buckets:
        try:
            supabase.storage.from_(b).remove([storage_path])
        except:
            pass


def build_output_filename(run_uuid: str, original_filename: str, ext: str = ".xlsx") -> str:
    if not ext.startswith("."):
        ext = f".{ext}"

    original_name = os.path.basename((original_filename or "").strip())
    original_stem, _ = os.path.splitext(original_name)
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", original_stem).strip("._")

    if safe_stem:
        return f"{run_uuid}_{safe_stem}{ext}"
    return f"{run_uuid}{ext}"


def get_bearer_token(authorization: str | None = Header(default=None, alias="Authorization")) -> str:
    if not authorization:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Missing Authorization header",
        )
    parts = authorization.strip().split(" ", 1)
    if len(parts) != 2 or parts[0].lower() != "bearer" or not parts[1].strip():
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid Authorization header",
        )
    return parts[1].strip()


def get_current_user_id(token: str = Depends(get_bearer_token)) -> str:
    try:
        auth_resp = supabase.auth.get_user(token)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid or expired token",
        )

    user = getattr(auth_resp, "user", None)
    if user is None and isinstance(auth_resp, dict):
        user = auth_resp.get("user")

    user_id = None
    if isinstance(user, dict):
        user_id = user.get("id")
    else:
        user_id = getattr(user, "id", None)

    if not user_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Unable to resolve user from token",
        )
    return user_id


def is_admin(user_id: str) -> bool:
    try:
        # Use a deterministic table lookup with service-role privileges.
        # This avoids ambiguous truthiness from RPC payload shapes.
        rows = (
            supabase.table("user_roles")
            .select("role")
            .eq("user_id", user_id)
            .eq("role", "admin")
            .limit(1)
            .execute()
            .data
        )
        return len(rows or []) > 0
    except Exception:
        return False


def require_run_access(run_id: str, user_id: str) -> dict:
    run_rows = (
        supabase.table("runs")
        .select("*")
        .eq("id", run_id)
        .limit(1)
        .execute()
        .data
    )
    if not run_rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Run not found")

    run = run_rows[0]
    run_owner_id = str(run.get("user_id") or "")
    caller_id = str(user_id or "")
    if run_owner_id != caller_id and not is_admin(caller_id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
    return run


def require_run_owner(run_id: str, user_id: str) -> dict:
    run_rows = (
        supabase.table("runs")
        .select("*")
        .eq("id", run_id)
        .limit(1)
        .execute()
        .data
    )
    if not run_rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Run not found")

    run = run_rows[0]
    if str(run.get("user_id") or "") != str(user_id or ""):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Forbidden")
    return run


def enforce_rate_limit(bucket: str, key: str):
    limit, window_sec = RATE_LIMITS[bucket]
    now = datetime.utcnow().timestamp()

    with _rate_lock:
        q = _rate_buckets[(bucket, key)]
        while q and (now - q[0]) > window_sec:
            q.popleft()
        if len(q) >= limit:
            raise HTTPException(
                status_code=status.HTTP_429_TOO_MANY_REQUESTS,
                detail="Rate limit exceeded",
            )
        q.append(now)



# =========================================================
# MAIN EXECUTION
# =========================================================

def execute_run(run_id: str):

    run_dir = os.path.join(BASE_WORKDIR, f"run_{run_id}")
    os.makedirs(run_dir, exist_ok=True)

    try:
        # ---------------------------------------
        # 1. Mark running
        # ---------------------------------------
        supabase.table("runs").update({
            "status": "running",
            "start_time": datetime.utcnow().isoformat()
        }).eq("id", run_id).execute()

        run = (
            supabase.table("runs")
            .select("*")
            .eq("id", run_id)
            .single()
            .execute()
            .data
        )

        log(run_id, "INFO", "Run started")

        op_path = run["op_filename"]
        ip_path = run["ip_filename"]
        master_path = run["master_filename"]

        if not op_path or not ip_path or not master_path:
            raise Exception("Missing input files in DB")

        # ---------------------------------------
        # 2. Prepare paths
        # ---------------------------------------
        op_local = os.path.join(run_dir, os.path.basename(op_path))
        ip_local = os.path.join(run_dir, os.path.basename(ip_path))
        master_local = os.path.join(run_dir, os.path.basename(master_path))
        output_local = os.path.join(run_dir, "output.xlsx")

        log(run_id, "INFO", f"OP file: {op_path}")
        log(run_id, "INFO", f"IP file: {ip_path}")
        log(run_id, "INFO", f"MASTER file: {master_path}")

        # ---------------------------------------
        # 3. Download
        # ---------------------------------------
        log(run_id, "INFO", "Downloading input files")

        download_from_storage("input-files", op_path, op_local)
        download_from_storage("crawl-input", ip_path, ip_local)
        download_from_storage("masters", master_path, master_local)

        log(run_id, "INFO", "All files downloaded successfully")

        # ---------------------------------------
        # 4. Run script (REALTIME LOGGING)
        # ---------------------------------------
        log(run_id, "INFO", "Starting Python script")

        process = subprocess.Popen(
            [
                "python",
                "-u",                     # 🔥 VERY IMPORTANT → unbuffered
                "pl_conso_check.py",
                op_local,
                ip_local,
                master_local,
                output_local,
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            bufsize=1
        )
        supabase.table("runs").update({
            "process_pid": process.pid
        }).eq("id", run_id).execute()


        # 🔥 STREAM LIVE LOGS
        for line in iter(process.stdout.readline, ''):
            log(run_id, "INFO", line.rstrip())

        process.stdout.close()
        process.wait()
        current_status = (
            supabase.table("runs")
            .select("status")
            .eq("id", run_id)
            .single()
            .execute()
            .data
        )
        if current_status and current_status["status"] == "cancelled":
            log(run_id, "INFO", "Run cancelled by user")
            return
        if process.returncode == -9:
            log(run_id, "INFO", "Run cancelled by user")
            return   # DO NOT go to except



        if process.returncode != 0:
            raise Exception("Script failed")

        log(run_id, "INFO", "Script finished successfully")

        # ---------------------------------------
        # 5. AI analysis
        # ---------------------------------------
        try:
            log(run_id, "INFO", "AI analysis started")
            log(run_id, "INFO", "Sending output to AI")

            ai_report = analyze_output_with_gemini(output_local)

            supabase.table("run_ai_reports").insert({
                "run_id": run_id,
                "report_json": ai_report,
                "summary": ai_report.get("summary"),
                "accuracy": ai_report.get("accuracy"),
                "verdict": ai_report.get("verdict")
            }).execute()

            log(run_id, "INFO", "AI report generated")
            log(run_id, "INFO", f"AI summary: {ai_report.get('summary', '')}")
            log(run_id, "INFO", f"AI accuracy: {ai_report.get('accuracy', '')}")
            log(run_id, "INFO", f"AI verdict: {ai_report.get('verdict', '')}")

        except Exception as e:
            log(run_id, "ERROR", f"AI failed: {str(e)}")

        # ---------------------------------------
        # 6. Upload result
        # ---------------------------------------
        filename = build_output_filename(run["run_uuid"], run.get("op_filename", ""), ".xlsx")

        upload_to_storage("run-outputs", filename, output_local)

        supabase.table("run_files").insert({
            "run_id": run_id,
            "filename": filename,
            "file_type": "FINAL_OUTPUT",
            "storage_path": filename
        }).execute()

        # -------------------------
        # 8. complete
        # -------------------------
        supabase.table("runs").update({
            "status": "completed",
            "end_time": datetime.utcnow().isoformat(),
            "process_pid": None
        }).eq("id", run_id).execute()

    except Exception as e:
        log(run_id, "ERROR", str(e))

        current = (
            supabase.table("runs")
            .select("status")
            .eq("id", run_id)
            .single()
            .execute()
            .data
        )

        # never override cancelled
        if not current or current["status"] != "cancelled":
            supabase.table("runs").update({
                "status": "failed",
                "end_time": datetime.utcnow().isoformat(),
                "process_pid": None
            }).eq("id", run_id).execute()

    finally:
        shutil.rmtree(run_dir, ignore_errors=True)

def execute_input_run(run_id: str):

    run_dir = os.path.join(BASE_WORKDIR, f"input_run_{run_id}")
    os.makedirs(run_dir, exist_ok=True)

    try:
        # -----------------------------
        # 1. mark running
        # -----------------------------
        supabase.table("runs").update({
            "status": "running",
            "start_time": datetime.utcnow().isoformat()
        }).eq("id", run_id).execute()

        run = (
            supabase.table("runs")
            .select("*")
            .eq("id", run_id)
            .single()
            .execute()
            .data
        )

        log(run_id, "INFO", "Input creation started")

        # -----------------------------
        # 2. local paths
        # -----------------------------
        biz_local = os.path.join(run_dir, "biz.xlsx")
        crawl_local = os.path.join(run_dir, "crawl.xlsx")
        template_local = os.path.join(run_dir, "template.xlsx")
        output_local = os.path.join(run_dir, "merged_output.xlsx")
        zip_local = os.path.join(run_dir, "all_tsv_files.zip")


        # -----------------------------
        # 3. download files
        # -----------------------------
        download_from_storage(
            "input-creation-bussiness-file",
            run["op_filename"],
            biz_local
        )

        download_from_storage(
            "input-creation-crawl-team-file",
            run["ip_filename"],
            crawl_local
        )

        download_from_storage(
            "input-creation-final-template",
            run["master_filename"],
            template_local
        )

        log(run_id, "INFO", "All files downloaded")

        # -----------------------------
        # 4. run python script
        # -----------------------------
        process = subprocess.Popen(
            [
                "python",
                "-u",
                "input_creation_script.py",
                biz_local,
                crawl_local,
                template_local,
                output_local
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True
        )
        supabase.table("runs").update({
            "process_pid": process.pid
        }).eq("id", run_id).execute()


        for line in process.stdout:
            log(run_id, "INFO", line.rstrip())

        process.wait()

        # 🔥 CANCELLED PROCESS (killed)
        if process.returncode == -9:
            log(run_id, "INFO", "Run cancelled by user")
            return

        # real failure
        if process.returncode != 0:
            raise Exception("Script failed")


        # -----------------------------
        # 5. upload result
        # -----------------------------
        output_filename = build_output_filename(run["run_uuid"], run.get("op_filename", ""), ".xlsx")

        upload_to_storage(
            "input-creation-output",
            output_filename,
            output_local
        )

        supabase.table("run_files").insert({
            "run_id": run_id,
            "filename": output_filename,
            "file_type": "MERGED_OUTPUT",
            "storage_path": output_filename
        }).execute()

        # -----------------------------
        # Upload TSV ZIP (ADD THIS)
        # -----------------------------
        zip_filename = f"{run['run_uuid']}_tsv.zip"

        upload_to_storage(
            "input-creation-output",
             zip_filename,
             zip_local
        )

        supabase.table("run_files").insert({
            "run_id": run_id,
            "filename": zip_filename,
            "file_type": "TSV_ZIP",
            "storage_path": zip_filename
        }).execute()


        # -----------------------------
        # 6. mark completed
        # -----------------------------
        supabase.table("runs").update({
            "status": "completed",
            "end_time": datetime.utcnow().isoformat()
        }).eq("id", run_id).execute()

        log(run_id, "INFO", "Input creation completed")

    except Exception as e:
        log(run_id, "ERROR", str(e))

        current = (
            supabase.table("runs")
            .select("status")
            .eq("id", run_id)
            .single()
            .execute()
            .data
        )

        if current and current["status"] == "cancelled":
            return

        supabase.table("runs").update({
            "status": "failed",
            "end_time": datetime.utcnow().isoformat()
        }).eq("id", run_id).execute()

    finally:
        shutil.rmtree(run_dir, ignore_errors=True)

def execute_pdp_run(run_id: str):

    run_dir = os.path.join(BASE_WORKDIR, f"pdp_run_{run_id}")
    os.makedirs(run_dir, exist_ok=True)

    try:
        # -----------------------------
        # 1. mark running
        # -----------------------------
        supabase.table("runs").update({
            "status": "running",
            "start_time": datetime.utcnow().isoformat()
        }).eq("id", run_id).execute()

        run = (
            supabase.table("runs")
            .select("*")
            .eq("id", run_id)
            .single()
            .execute()
            .data
        )

        log(run_id, "INFO", "PDP check started")

        # -----------------------------
        # 2. local paths
        # -----------------------------
        op_local = os.path.join(run_dir, "pdp_output.tsv")
        ip_local = os.path.join(run_dir, "pdp_input.tsv")
        master_local = os.path.join(run_dir, "pdp_master.csv")
        output_local = os.path.join(run_dir, "pdp_output.xlsx")

        # -----------------------------
        # 3. download files
        # -----------------------------
        log(run_id, "INFO", f"OP file: {run['op_filename']}")
        log(run_id, "INFO", f"IP file: {run['ip_filename']}")
        log(run_id, "INFO", f"MASTER file: {run['master_filename']}")

        if not run.get("op_filename") or not run.get("ip_filename") or not run.get("master_filename"):
            raise Exception("Missing input files in DB")

        log(run_id, "INFO", "Downloading input files")

        download_from_storage(
            "pdp-input",
            run["op_filename"],
            op_local
        )

        download_from_storage(
            "pdp-crawl-input",
            run["ip_filename"],
            ip_local
        )

        download_from_storage(
            "pdp-masters",
            run["master_filename"],
            master_local
        )

        # -----------------------------
        # 3b. cache lookup for input counts (by SHA256)
        # -----------------------------
        cache_bucket = "pdp-cache"
        cache_hit_path = None
        cache_key = None
        try:
            cache_key = f"{sha256_file(ip_local)}.json"
            cache_local = os.path.join(run_dir, "pdp_ip_counts_cache.json")
            log(run_id, "INFO", f"Looking for input counts cache: {cache_key}")
            download_from_storage(cache_bucket, cache_key, cache_local)
            cache_hit_path = cache_local
            log(run_id, "INFO", "Input counts cache hit")
        except Exception as e:
            log(run_id, "INFO", f"No input counts cache (will compute). {str(e)}")

        # If input is Excel, convert locally to CSV for faster reads (do not upload)
        ip_filename = run.get("ip_filename", "") or ""
        if ip_filename.lower().endswith((".xlsx", ".xls")) and os.getenv("PDP_CONVERT_INPUT", "0").strip() == "1":
            try:
                import csv
                from openpyxl import load_workbook
                import shutil as _shutil

                ip_csv_local = os.path.join(run_dir, "pdp_input.csv")
                log(run_id, "INFO", "Converting input Excel to CSV (local, streaming)")

                # If local path doesn't have an excel extension, copy to a temp .xlsx
                excel_path = ip_local
                if not ip_local.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")):
                    excel_path = os.path.join(run_dir, "pdp_input.xlsx")
                    _shutil.copyfile(ip_local, excel_path)

                wb = load_workbook(excel_path, read_only=True, data_only=True)
                ws = wb.active

                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                header = [str(v).strip().lower() if v is not None else "" for v in header_row]

                scope_idx = header.index("scope") if "scope" in header else (header.index("scope_name") if "scope_name" in header else None)
                rname_idx = header.index("rname") if "rname" in header else (header.index("domain_input") if "domain_input" in header else None)

                if scope_idx is None or rname_idx is None:
                    raise Exception("Required columns not found in input Excel")

                max_col = max(scope_idx, rname_idx) + 1
                with open(ip_csv_local, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(["scope_name", "domain_input"])
                    for row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
                        scope_val = row[scope_idx]
                        rname_val = row[rname_idx]
                        writer.writerow([
                            "" if scope_val is None else str(scope_val),
                            "" if rname_val is None else str(rname_val)
                        ])

                wb.close()
                ip_local = ip_csv_local
                log(run_id, "INFO", "Input Excel converted to CSV (local, streaming)")
            except Exception as e:
                log(run_id, "ERROR", f"Excel to CSV conversion failed, using original file. {str(e)}")

        log(run_id, "INFO", "All files downloaded")

        # -----------------------------
        # 4. run python script
        # -----------------------------
        process = subprocess.Popen(
            [
                "python",
                "-u",
                "pdp_check.py",
                op_local,
                ip_local,
                master_local,
                output_local
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            env={
                **os.environ,
                "PDP_OP_NAME": run.get("op_filename", ""),
                "PDP_IP_NAME": run.get("ip_filename", ""),
                "PDP_MASTER_NAME": run.get("master_filename", ""),
                "PDP_IP_COUNTS_CACHE": cache_hit_path or "",
                "PDP_IP_COUNTS_CACHE_WRITE": os.path.join(run_dir, "pdp_ip_counts_cache.json") if cache_hit_path is None else ""
            }
        )
        supabase.table("runs").update({
            "process_pid": process.pid
        }).eq("id", run_id).execute()

        for line in process.stdout:
            log(run_id, "INFO", line.rstrip())

        process.wait()

        if process.returncode == -9:
            log(run_id, "INFO", "Run cancelled by user")
            return

        if process.returncode != 0:
            raise Exception("Script failed")

        # -----------------------------
        # 4b. upload cache if computed
        # -----------------------------
        try:
            if cache_key and cache_hit_path is None:
                cache_local = os.path.join(run_dir, "pdp_ip_counts_cache.json")
                if os.path.exists(cache_local) and os.path.getsize(cache_local) > 0:
                    upload_to_storage(cache_bucket, cache_key, cache_local)
                    log(run_id, "INFO", f"Input counts cache uploaded: {cache_key}")
        except Exception as e:
            log(run_id, "ERROR", f"Failed to upload input counts cache: {str(e)}")

        # -----------------------------
        # 5. AI analysis
        # -----------------------------
        try:
            log(run_id, "INFO", "AI analysis started")
            log(run_id, "INFO", "Sending output to AI")

            ai_report = analyze_output_with_gemini(output_local)

            supabase.table("run_ai_reports").insert({
                "run_id": run_id,
                "report_json": ai_report,
                "summary": ai_report.get("summary"),
                "accuracy": ai_report.get("accuracy"),
                "verdict": ai_report.get("verdict")
            }).execute()

            log(run_id, "INFO", "AI report generated")
            log(run_id, "INFO", f"AI summary: {ai_report.get('summary', '')}")
            log(run_id, "INFO", f"AI accuracy: {ai_report.get('accuracy', '')}")
            log(run_id, "INFO", f"AI verdict: {ai_report.get('verdict', '')}")

        except Exception as e:
            log(run_id, "ERROR", f"AI failed: {str(e)}")

        # -----------------------------
        # 6. upload result
        # -----------------------------
        output_filename = build_output_filename(run["run_uuid"], run.get("op_filename", ""), ".xlsx")

        upload_to_storage(
            "pdp-run-output",
            output_filename,
            output_local
        )

        supabase.table("run_files").insert({
            "run_id": run_id,
            "filename": output_filename,
            "file_type": "FINAL_OUTPUT",
            "storage_path": output_filename
        }).execute()

        # -----------------------------
        # 7. mark completed
        # -----------------------------
        supabase.table("runs").update({
            "status": "completed",
            "end_time": datetime.utcnow().isoformat(),
            "process_pid": None
        }).eq("id", run_id).execute()

        log(run_id, "INFO", "PDP check completed")

    except Exception as e:
        log(run_id, "ERROR", str(e))

        current = (
            supabase.table("runs")
            .select("status")
            .eq("id", run_id)
            .single()
            .execute()
            .data
        )

        if current and current["status"] == "cancelled":
            return

        supabase.table("runs").update({
            "status": "failed",
            "end_time": datetime.utcnow().isoformat(),
            "process_pid": None
        }).eq("id", run_id).execute()

    finally:
        shutil.rmtree(run_dir, ignore_errors=True)


def execute_pp_run(run_id: str):

    run_dir = os.path.join(BASE_WORKDIR, f"pp_run_{run_id}")
    os.makedirs(run_dir, exist_ok=True)

    try:
        # -----------------------------
        # 1. mark running
        # -----------------------------
        supabase.table("runs").update({
            "status": "running",
            "start_time": datetime.utcnow().isoformat()
        }).eq("id", run_id).execute()

        run = (
            supabase.table("runs")
            .select("*")
            .eq("id", run_id)
            .single()
            .execute()
            .data
        )

        log(run_id, "INFO", "AE PP check started")

        # -----------------------------
        # 2. local paths
        # -----------------------------
        op_name = run.get("op_filename")
        ip_name = run.get("ip_filename")
        master_name = run.get("master_filename")
        ae_name = run.get("ae_filename")

        op_local = os.path.join(run_dir, os.path.basename(op_name)) if op_name else ""
        ip_local = os.path.join(run_dir, os.path.basename(ip_name)) if ip_name else ""
        master_local = os.path.join(run_dir, os.path.basename(master_name))
        ae_template_local = os.path.join(run_dir, os.path.basename(ae_name)) if ae_name else ""
        output_local = os.path.join(run_dir, "pp_output.xlsx")

        log(run_id, "INFO", f"OP file: {op_name}")
        log(run_id, "INFO", f"IP file: {ip_name}")
        log(run_id, "INFO", f"MASTER file: {master_name}")
        log(run_id, "INFO", f"AE template file: {ae_name}")

        if not master_name:
            raise Exception("Missing reference/master file in DB")
        if not op_name and not ip_name:
            raise Exception("Missing both source and review files in DB")

        # -----------------------------
        # 3. download files
        # -----------------------------
        log(run_id, "INFO", "Downloading input files")

        if op_name and op_local:
            download_from_storage("pp-input", op_name, op_local)
        if ip_name and ip_local:
            download_from_storage("pp-review-input", ip_name, ip_local)
        download_from_storage("pp-reference", master_name, master_local)
        if ae_name and ae_template_local:
            download_from_storage("pp-ae-checks", ae_name, ae_template_local)

        log(run_id, "INFO", "All files downloaded")

        # -----------------------------
        # 4. run python script
        # -----------------------------
        process = subprocess.Popen(
            [
                "python",
                "-u",
                "PP_conso_check 1.py",
                op_local if op_name else "",
                ip_local if ip_name else "",
                master_local,
                output_local,
                ae_template_local if ae_name else ""
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True
        )
        supabase.table("runs").update({
            "process_pid": process.pid
        }).eq("id", run_id).execute()

        for line in process.stdout:
            log(run_id, "INFO", line.rstrip())

        process.wait()

        if process.returncode == -9:
            log(run_id, "INFO", "Run cancelled by user")
            return

        if process.returncode != 0:
            raise Exception("Script failed")

        # -----------------------------
        # 5. upload result
        # -----------------------------
        base_name = run.get("op_filename") or run.get("ip_filename") or "pp_output.xlsx"
        output_filename = build_output_filename(run["run_uuid"], base_name, ".xlsx")

        upload_to_storage(
            "pp-run-output",
            output_filename,
            output_local
        )

        supabase.table("run_files").insert({
            "run_id": run_id,
            "filename": output_filename,
            "file_type": "FINAL_OUTPUT",
            "storage_path": output_filename
        }).execute()

        # -----------------------------
        # 6. mark completed
        # -----------------------------
        supabase.table("runs").update({
            "status": "completed",
            "end_time": datetime.utcnow().isoformat(),
            "process_pid": None
        }).eq("id", run_id).execute()

        log(run_id, "INFO", "AE PP check completed")

    except Exception as e:
        log(run_id, "ERROR", str(e))

        current = (
            supabase.table("runs")
            .select("status")
            .eq("id", run_id)
            .single()
            .execute()
            .data
        )

        if current and current["status"] == "cancelled":
            return

        supabase.table("runs").update({
            "status": "failed",
            "end_time": datetime.utcnow().isoformat(),
            "process_pid": None
        }).eq("id", run_id).execute()

    finally:
        shutil.rmtree(run_dir, ignore_errors=True)


# =========================================================
# API
# =========================================================

@app.post("/run/{run_id}")
def start_run(run_id: str, bg: BackgroundTasks, request: Request, user_id: str = Depends(get_current_user_id)):
    enforce_rate_limit("trigger", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    require_run_access(run_id, user_id)
    bg.add_task(execute_run, run_id)
    return {"status": "started"}


@app.post("/run/{run_id}/rerun")
def rerun(run_id: str, bg: BackgroundTasks, request: Request, user_id: str = Depends(get_current_user_id)):
    enforce_rate_limit("trigger", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    old = require_run_access(run_id, user_id)

    new = (
        supabase.table("runs")
        .insert({
            "user_id": old["user_id"],
            "project_id": old["project_id"],
            "site_id": old["site_id"],
            "scope": old["scope"],
            "op_filename": old["op_filename"],
            "ip_filename": old["ip_filename"],
            "master_filename": old["master_filename"],
            "status": "pending",
             "automation_slug": old["automation_slug"]
        })
        .execute()
        .data[0]
    )

    bg.add_task(execute_run, new["id"])

    return {"status": "rerun_started"}


@app.get("/run/{run_id}/ai-report")
def get_ai_report(run_id: str, user_id: str = Depends(get_current_user_id)):
    require_run_owner(run_id, user_id)
    return (
        supabase.table("run_ai_reports")
        .select("*")
        .eq("run_id", run_id)
        .single()
        .execute()
        .data
    )


@app.get("/run/{run_id}/ai-report-pdf")
def download_ai_report_pdf(run_id: str, user_id: str = Depends(get_current_user_id)):
    require_run_owner(run_id, user_id)

    row = (
        supabase.table("run_ai_reports")
        .select("*")
        .eq("run_id", run_id)
        .single()
        .execute()
        .data
    )

    tmp = f"/tmp/{run_id}.pdf"

    generate_pdf_from_ai_report(row["report_json"], tmp)

    return FileResponse(tmp, media_type="application/pdf")

@app.get("/run/{run_id}/logs")
def get_run_logs(
    run_id: str,
    request: Request,
    since_id: int | None = None,
    user_id: str = Depends(get_current_user_id),
):
    """
    Return logs in insertion order. Optional since_id enables real-time polling.
    """
    enforce_rate_limit("logs", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    require_run_owner(run_id, user_id)
    q = (
        supabase.table("run_logs")
        .select("*")
        .eq("run_id", run_id)
    )
    if since_id is not None:
        q = q.gt("id", since_id)
    return (
        q.order("id", desc=False)
        .execute()
        .data
    )

@app.post("/input-run/{run_id}")
def start_input_run(run_id: str, bg: BackgroundTasks, request: Request, user_id: str = Depends(get_current_user_id)):
    enforce_rate_limit("trigger", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    require_run_access(run_id, user_id)
    bg.add_task(execute_input_run, run_id)
    return {"status": "started"}

@app.post("/input-run/{run_id}/rerun")
def rerun_input(run_id: str, bg: BackgroundTasks, request: Request, user_id: str = Depends(get_current_user_id)):
    enforce_rate_limit("trigger", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    old = require_run_access(run_id, user_id)

    # create new run row
    new = (
        supabase.table("runs")
        .insert({
            "user_id": old["user_id"],
            "project_id": old["project_id"],
            "site_id": old["site_id"],
            "scope": old["scope"],
            "op_filename": old["op_filename"],
            "ip_filename": old["ip_filename"],
            "master_filename": old["master_filename"],
            "status": "pending",
            "automation_slug": old["automation_slug"]

        })
        .execute()
        .data[0]
    )

    # start INPUT execution
    bg.add_task(execute_input_run, new["id"])

    return {"status": "input_rerun_started"}

@app.post("/pdp-run/{run_id}")
def start_pdp_run(run_id: str, bg: BackgroundTasks, request: Request, user_id: str = Depends(get_current_user_id)):
    enforce_rate_limit("trigger", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    require_run_access(run_id, user_id)
    bg.add_task(execute_pdp_run, run_id)
    return {"status": "started"}

@app.post("/pdp-run/{run_id}/rerun")
def rerun_pdp(run_id: str, bg: BackgroundTasks, request: Request, user_id: str = Depends(get_current_user_id)):
    enforce_rate_limit("trigger", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    old = require_run_access(run_id, user_id)

    new = (
        supabase.table("runs")
        .insert({
            "user_id": old["user_id"],
            "project_id": old["project_id"],
            "site_id": old["site_id"],
            "scope": old["scope"],
            "op_filename": old["op_filename"],
            "ip_filename": old["ip_filename"],
            "ae_filename": old.get("ae_filename"),
            "master_filename": old["master_filename"],
            "status": "pending",
            "automation_slug": old["automation_slug"]
        })
        .execute()
        .data[0]
    )

    bg.add_task(execute_pdp_run, new["id"])

    return {"status": "pdp_rerun_started"}


@app.post("/pp-run/{run_id}")
def start_pp_run(run_id: str, bg: BackgroundTasks, request: Request, user_id: str = Depends(get_current_user_id)):
    enforce_rate_limit("trigger", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    require_run_access(run_id, user_id)
    bg.add_task(execute_pp_run, run_id)
    return {"status": "started"}


@app.post("/pp-run/{run_id}/rerun")
def rerun_pp(run_id: str, bg: BackgroundTasks, request: Request, user_id: str = Depends(get_current_user_id)):
    enforce_rate_limit("trigger", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    old = require_run_access(run_id, user_id)

    new = (
        supabase.table("runs")
        .insert({
            "user_id": old["user_id"],
            "project_id": old["project_id"],
            "site_id": old["site_id"],
            "scope": old["scope"],
            "op_filename": old["op_filename"],
            "ip_filename": old["ip_filename"],
            "master_filename": old["master_filename"],
            "status": "pending",
            "automation_slug": old["automation_slug"]
        })
        .execute()
        .data[0]
    )

    bg.add_task(execute_pp_run, new["id"])

    return {"status": "pp_rerun_started"}

# =========================================================
# UNIVERSAL DELETE RUN (PL CONSO + PL INPUT)na
# =========================================================


@app.post("/runs/{run_id}/cancel")
def cancel_run(run_id: str, request: Request, user_id: str = Depends(get_current_user_id)):
    enforce_rate_limit("mutating_run", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    run = require_run_access(run_id, user_id)

    # -------------------------
    # 1. HARD KILL PROCESS
    # -------------------------
    pid = run.get("process_pid")

    if pid:
        try:
            os.kill(pid, signal.SIGKILL)
        except:
            pass

    # -------------------------
    # 2. DELETE STORAGE FILES
    # -------------------------
    files = (
        supabase.table("run_files")
        .select("*")
        .eq("run_id", run_id)
        .execute()
        .data
    )

    for f in files or []:
        remove_from_buckets(
            f["storage_path"],
            ["run-outputs", "input-creation-output", "pdp-run-output", "pp-run-output"]
        )

    # -------------------------
    # 3. CLEAN CHILD TABLES
    # -------------------------
    supabase.table("run_logs").delete().eq("run_id", run_id).execute()
    supabase.table("run_files").delete().eq("run_id", run_id).execute()
    supabase.table("run_ai_reports").delete().eq("run_id", run_id).execute()

    # -------------------------
    # 4. UPDATE STATUS ONLY
    # -------------------------
    supabase.table("runs").update({
        "status": "cancelled",
        "end_time": datetime.utcnow().isoformat(),
        "process_pid": None
    }).eq("id", run_id).execute()

    return {"status": "cancelled"}

@app.post("/runs/{run_id}/delete")
def delete_run(run_id: str, request: Request, user_id: str = Depends(get_current_user_id)):
    enforce_rate_limit("mutating_run", f"{user_id}:{request.client.host if request.client else 'unknown'}")
    run = require_run_access(run_id, user_id)

    files = (
        supabase.table("run_files")
        .select("*")
        .eq("run_id", run_id)
        .execute()
        .data
    )

    for f in files or []:
        remove_from_buckets(
            f["storage_path"],
            ["run-outputs", "input-creation-output", "pdp-run-output", "pp-run-output"]
        )

    supabase.table("run_logs").delete().eq("run_id", run_id).execute()
    supabase.table("run_files").delete().eq("run_id", run_id).execute()
    supabase.table("run_ai_reports").delete().eq("run_id", run_id).execute()

    supabase.table("runs").delete().eq("id", run_id).execute()

    return {"status": "deleted"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
