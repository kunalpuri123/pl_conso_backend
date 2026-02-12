import pandas as pd
import numpy as np
import re
from pathlib import Path
import sys
from datetime import datetime
import os
import time
from openpyxl import load_workbook
import json

print("=== PDP CHECKLIST STARTED ===", flush=True)

# ================= ARGUMENTS =================
if len(sys.argv) != 5:
    print("Usage: python pdp_check.py <CRAWL_OUTPUT> <CRAWL_INPUT> <MASTER_FILE> <OUTPUT_FILE>")
    sys.exit(1)

OP_FILE = Path(sys.argv[1])   # crawl output to be checked
IP_FILE = Path(sys.argv[2])   # crawl input (consolidated)
MASTER_FILE = Path(sys.argv[3])
OUTPUT_FILE = Path(sys.argv[4])

DISPLAY_NAMES = {
    str(OP_FILE): os.environ.get("PDP_OP_NAME", ""),
    str(IP_FILE): os.environ.get("PDP_IP_NAME", ""),
    str(MASTER_FILE): os.environ.get("PDP_MASTER_NAME", ""),
}

def display_name(fp: Path) -> str:
    return DISPLAY_NAMES.get(str(fp), "") or fp.name

print(f"📄 Output File : {display_name(OP_FILE)}")
print(f"📄 Input File  : {display_name(IP_FILE)}")
print(f"📄 Master File : {display_name(MASTER_FILE)}")
print(f"📄 Result File : {OUTPUT_FILE}")

for fp, label in [(OP_FILE, "Output"), (IP_FILE, "Input"), (MASTER_FILE, "Master")]:
    if not fp.exists():
        print(f"❌ {label} file not found: {fp}")
        sys.exit(1)

def is_excel_display(fp: Path) -> bool:
    return display_name(fp).lower().endswith((".xlsx", ".xls"))

def is_excel_file(fp: Path) -> bool:
    try:
        with open(fp, "rb") as f:
            sig = f.read(8)
        # XLSX is a zip archive (PK\x03\x04). XLS (OLE) starts with D0 CF 11 E0 A1 B1 1A E1
        return sig.startswith(b"PK\x03\x04") or sig.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1")
    except Exception:
        return False

def excel_compatible_path(fp: Path) -> Path | None:
    """
    Ensure a path with an Excel-compatible extension if the content is Excel.
    Returns a temp path if created, or the original path if already compatible.
    """
    excel_exts = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
    if fp.suffix.lower() in excel_exts:
        return fp
    if not is_excel_file(fp):
        return None
    tmp_path = fp.with_suffix(".xlsx")
    try:
        with open(fp, "rb") as src, open(tmp_path, "wb") as dst:
            dst.write(src.read())
        return tmp_path
    except Exception:
        return None

# ================= FILE HELPERS =================
# Tolerant TSV reader that keeps all rows by padding/truncating to header length
def read_tsv_tolerant(fp: Path, encoding: str = "utf-8") -> pd.DataFrame:
    with open(fp, "r", encoding=encoding, errors="replace") as f:
        header_line = f.readline().rstrip("\n")
        headers = header_line.split("\t")
        rows = []
        for line in f:
            line = line.rstrip("\n")
            parts = line.split("\t")
            if len(parts) < len(headers):
                parts.extend([""] * (len(headers) - len(parts)))
            elif len(parts) > len(headers):
                # merge extras into last column to preserve data
                parts = parts[: len(headers) - 1] + ["\t".join(parts[len(headers) - 1 :])]
            rows.append(parts)
    return pd.DataFrame(rows, columns=headers)

def read_file(fp: Path) -> pd.DataFrame:
    display = display_name(fp)
    display_lower = display.lower()
    suffix = fp.suffix.lower()

    # If original filename indicates Excel, prefer read_excel only if content is Excel
    if display_lower.endswith((".xlsx", ".xls")):
        excel_path = excel_compatible_path(fp)
        if excel_path is None:
            pass
        else:
            # If this is the input file, load only needed columns to speed up
            if fp == IP_FILE:
                print(f"⏳ Loading Excel (header only): {display}", flush=True)
                header_df = pd.read_excel(excel_path, nrows=0)
                header_cols = [c.strip().lower() for c in header_df.columns]
                scope_col = "scope" if "scope" in header_cols else ("scope_name" if "scope_name" in header_cols else None)
                if "rname" in header_cols:
                    rname_col = "rname"
                elif "site_name" in header_cols:
                    rname_col = "site_name"
                else:
                    rname_col = "domain_input" if "domain_input" in header_cols else None

                if scope_col and rname_col:
                    print(f"⏳ Loading Excel (usecols={scope_col},{rname_col}): {display}", flush=True)
                    return pd.read_excel(excel_path, dtype=str, keep_default_na=False, usecols=[scope_col, rname_col])

            print(f"⏳ Loading Excel: {display}", flush=True)
            return pd.read_excel(excel_path, dtype=str, keep_default_na=False)

    if suffix == ".tsv" or display_lower.endswith(".tsv"):
        try:
            print(f"⏳ Loading TSV: {display}", flush=True)
            return pd.read_csv(fp, sep="\t", dtype=str, encoding="utf-8", keep_default_na=False)
        except UnicodeDecodeError:
            print(f"⚠️ WARNING: {display} is not UTF-8. Retrying with latin1 tolerant parser (no row drops).")
            return read_tsv_tolerant(fp, encoding="latin1")
        except pd.errors.ParserError:
            print(f"⚠️ WARNING: {display} has malformed lines. Retrying with tolerant TSV parser (no row drops).")
            return read_tsv_tolerant(fp, encoding="utf-8")
    if suffix == ".csv" or display_lower.endswith(".csv"):
        print(f"⏳ Loading CSV: {display}", flush=True)
        return pd.read_csv(fp, dtype=str, keep_default_na=False)
    print(f"⏳ Loading file: {display}", flush=True)
    return pd.read_excel(fp, dtype=str, keep_default_na=False)

def compute_ip_counts_excel(fp: Path, scopes_in_output: set | None, collect_input_map: bool = False):
    display = display_name(fp)
    print(f"⏳ Streaming Excel for counts: {display}", flush=True)
    wb = load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(v).strip().lower() if v is not None else "" for v in header_row]

    scope_idx = header.index("scope") if "scope" in header else (header.index("scope_name") if "scope_name" in header else None)
    if "rname" in header:
        rname_idx = header.index("rname")
    elif "site_name" in header:
        rname_idx = header.index("site_name")
    else:
        rname_idx = header.index("domain_input") if "domain_input" in header else None

    if scope_idx is None or rname_idx is None:
        wb.close()
        return None, None, None, None

    counts = {}
    input_map = {}
    total_rows = 0
    kept_rows = 0

    # Try to capture optional fields for input map
    base_id_idx = header.index("base_id") if "base_id" in header else None
    country_idx = header.index("country") if "country" in header else None
    pname_idx = header.index("pname") if "pname" in header else None
    skuvar_idx = header.index("skuvarient") if "skuvarient" in header else None

    max_col = max(i for i in [scope_idx, rname_idx, base_id_idx, country_idx, pname_idx, skuvar_idx] if i is not None) + 1
    for row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
        total_rows += 1
        scope_val = row[scope_idx]
        if scope_val is None:
            continue
        scope_s = str(scope_val).strip()
        if scopes_in_output and scope_s not in scopes_in_output:
            continue
        rname_val = row[rname_idx]
        rname_s = "" if rname_val is None else str(rname_val).strip()
        rname_s = normalize_rname_for_scope(scope_key_main, rname_s)
        counts[(scope_s, rname_s)] = counts.get((scope_s, rname_s), 0) + 1
        # build input map if columns exist
        if collect_input_map and base_id_idx is not None and country_idx is not None and (pname_idx is not None or skuvar_idx is not None):
            base_id = "" if row[base_id_idx] is None else str(row[base_id_idx]).strip()
            country = "" if row[country_idx] is None else str(row[country_idx]).strip()
            key = f"{rname_s}{base_id}{country}"
            pname_val = "" if pname_idx is None or row[pname_idx] is None else str(row[pname_idx]).strip()
            skuvar_val = "" if skuvar_idx is None or row[skuvar_idx] is None else str(row[skuvar_idx]).strip()
            if key not in input_map:
                input_map[key] = {"pname": pname_val, "skuvarient": skuvar_val}
        kept_rows += 1

    wb.close()
    return counts, total_rows, kept_rows, (header[scope_idx], header[rname_idx]), input_map

def compute_ip_counts_pandas_excel(fp: Path, scopes_in_output: set | None, collect_input_map: bool = False):
    display = display_name(fp)
    print(f"⏳ Loading Excel (pandas usecols): {display}", flush=True)
    header_df = pd.read_excel(fp, nrows=0)
    header_cols = [c.strip().lower() for c in header_df.columns]
    scope_col = "scope" if "scope" in header_cols else ("scope_name" if "scope_name" in header_cols else None)
    if "rname" in header_cols:
        rname_col = "rname"
    elif "site_name" in header_cols:
        rname_col = "site_name"
    else:
        rname_col = "domain_input" if "domain_input" in header_cols else None

    if scope_col is None or rname_col is None:
        return None, None, None, None, None

    # try to pull optional fields for input map
    base_id_col = "base_id" if "base_id" in header_cols else None
    country_col = "country" if "country" in header_cols else None
    pname_col = "pname" if "pname" in header_cols else None
    skuvar_col = "skuvarient" if "skuvarient" in header_cols else None
    extra_cols = [base_id_col, country_col, pname_col, skuvar_col] if collect_input_map else []
    usecols = [c for c in [scope_col, rname_col, *extra_cols] if c]

    df_ip = pd.read_excel(fp, dtype=str, keep_default_na=False, usecols=usecols)
    total_rows = len(df_ip)
    df_ip.columns = [c.strip().lower() for c in df_ip.columns]
    if "scope" not in df_ip.columns and "scope_name" in df_ip.columns:
        df_ip = df_ip.rename(columns={"scope_name": "scope"})
    if "rname" not in df_ip.columns and "domain_input" in df_ip.columns:
        df_ip = df_ip.rename(columns={"domain_input": "rname"})

    if scopes_in_output:
        df_ip = df_ip[df_ip["scope"].isin(scopes_in_output)]
    kept_rows = len(df_ip)
    ip_key = list(zip(
        df_ip["scope"].astype(str).str.strip(),
        df_ip["rname"].astype(str).str.strip().apply(lambda x: normalize_rname_for_scope(scope_key_main, x))
    ))
    counts = pd.Series(ip_key).value_counts().to_dict()

    input_map = {}
    if collect_input_map and {"base_id", "country"}.issubset(df_ip.columns):
        rname_s = df_ip["rname"].astype(str).str.strip()
        base_s = df_ip["base_id"].astype(str).str.strip()
        country_s = df_ip["country"].astype(str).str.strip()
        pname_s = df_ip["pname"].astype(str).str.strip() if "pname" in df_ip.columns else pd.Series([""] * len(df_ip))
        skuvar_s = df_ip["skuvarient"].astype(str).str.strip() if "skuvarient" in df_ip.columns else pd.Series([""] * len(df_ip))
        for rn, b, c, pn, sv in zip(rname_s, base_s, country_s, pname_s, skuvar_s):
            key = f"{rn}{b}{c}"
            if key not in input_map:
                input_map[key] = {"pname": pn, "skuvarient": sv}

    return counts, total_rows, kept_rows, (scope_col, rname_col), input_map

def encode_counts(counts: dict) -> dict:
    return {f"{k[0]}|||{k[1]}": v for k, v in counts.items()}

def decode_counts(data: dict) -> dict:
    counts = {}
    for k, v in data.items():
        if "|||" in k:
            scope, rname = k.split("|||", 1)
        else:
            scope, rname = k, ""
        counts[(scope, rname)] = int(v)
    return counts

def encode_input_map(input_map: dict) -> dict:
    return input_map

def decode_input_map(data: dict) -> dict:
    return data or {}

def is_na_text(val: str) -> bool:
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in {"", "n/a", "na", "null", "nan"}

def is_not_available(val: str) -> bool:
    if val is None:
        return False
    return str(val).strip().lower() == "not available"

def series_is_na(s: pd.Series) -> pd.Series:
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0]
    return s.isna() | s.astype(str).str.strip().str.lower().isin({"", "n/a", "na", "null", "nan"})

# ================= LOAD FILES =================
t0 = time.perf_counter()
df = read_file(OP_FILE)
print(f"✅ Loaded OP in {time.perf_counter() - t0:.2f}s | rows={len(df)}", flush=True)

t0 = time.perf_counter()
master_df = read_file(MASTER_FILE)
print(f"✅ Loaded MASTER in {time.perf_counter() - t0:.2f}s | rows={len(master_df)}", flush=True)

# Normalize column names
df.columns = [c.strip().lower() for c in df.columns]
master_df.columns = [c.strip().lower() for c in master_df.columns]

def scope_key_from_value(val: str) -> str:
    s = str(val or "").strip()
    if not s:
        return ""
    s = s.split("_")[0].strip()
    s = s.split(" ")[0].strip()
    alias = {"bnc": "benefit"}
    if s.lower() in alias:
        s = alias[s.lower()]
    return s.lower()

# Normalize retailer names for specific scopes (kept minimal)
def normalize_rname_for_scope(scope_key: str, name: str) -> str:
    n = str(name).strip()
    if scope_key == "mfk" and n.lower() == "nordstrom":
        return "Nordstrom"
    if scope_key == "hermes":
        # normalize to base retailer name: take prefix before underscore and remove spaces
        base = n.split("_")[0].replace(" ", "")
        return base.lower()
    return n

# Read input header without full load
def get_file_columns(fp: Path, display_override: str | None = None) -> list[str]:
    display = display_override or display_name(fp)
    display_lower = display.lower()
    try:
        if display_lower.endswith((".xlsx", ".xls")):
            excel_path = excel_compatible_path(fp)
            if excel_path is not None:
                wb = load_workbook(excel_path, read_only=True, data_only=True)
                ws = wb.active
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                wb.close()
                return [str(v).strip() if v is not None else "" for v in header_row]
        if display_lower.endswith(".tsv"):
            with open(fp, "r", encoding="utf-8", errors="ignore") as f:
                header = f.readline().rstrip("\n")
            return [c.strip() for c in header.split("\t")]
        if display_lower.endswith(".csv"):
            with open(fp, "r", encoding="utf-8", errors="ignore") as f:
                header = f.readline().rstrip("\n")
            return [c.strip() for c in header.split(",")]
    except Exception:
        pass
    return []

# Safe column access (handles duplicate column names)
def col_series(frame: pd.DataFrame, name: str, default: str = "") -> pd.Series:
    if name not in frame.columns:
        return pd.Series([default] * len(frame))
    col = frame[name]
    if isinstance(col, pd.DataFrame):
        col = col.iloc[:, 0]
    return col

# Benefit-specific column aliases (apply to all checks)
benefit_scope = any(scope_key_from_value(s) == "benefit" for s in df.get("scope", pd.Series([""])).astype(str).str.strip())
if benefit_scope:
    if "rname" not in df.columns and "site_name" in df.columns:
        df["rname"] = df["site_name"]
    if "rating" not in df.columns and "average_rating" in df.columns:
        df["rating"] = df["average_rating"]
    if "country" not in df.columns and "region_site" in df.columns:
        df["country"] = df["region_site"]
    if "regularprice" not in df.columns and "regular_price" in df.columns:
        df["regularprice"] = df["regular_price"]
    if "finalprice" not in df.columns and "final_price" in df.columns:
        df["finalprice"] = df["final_price"]

# ================= REQUIRED BASE COLUMNS =================
required_cols = ["scope", "rname", "country"]
for col in required_cols:
    if col not in df.columns:
        print(f"❌ Output file missing '{col}' column")
        sys.exit(1)

base_columns = list(df.columns)

scopes_in_output = set(df["scope"].dropna().astype(str).str.strip())
print("🔎 Scopes found in output:", scopes_in_output)

# Primary scope key (prefix before first underscore)
scope_key_main = ""
for s in scopes_in_output:
    scope_key_main = scope_key_from_value(s)
    if scope_key_main:
        break
if not scope_key_main and "scope" in master_df.columns:
    ms = master_df["scope"]
    if isinstance(ms, pd.DataFrame):
        ms = ms.iloc[:, 0]
    ms = ms.dropna().astype(str).str.strip()
    if len(ms):
        scope_key_main = scope_key_from_value(ms.iloc[0])

# ================= LOAD INPUT (FAST PATH FOR EXCEL) =================
CACHE_VERSION = 2
ip_df = None
ip_counts_precomputed = None
ip_input_map = {}
cache_loaded = False
cache_path = os.getenv("PDP_IP_COUNTS_CACHE", "").strip()
cache_write = os.getenv("PDP_IP_COUNTS_CACHE_WRITE", "").strip()
want_full_cache = bool(cache_write)
if cache_path and os.path.exists(cache_path):
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        if payload.get("version") == CACHE_VERSION:
            ip_counts_precomputed = decode_counts(payload.get("counts", {}))
            ip_input_map = decode_input_map(payload.get("input_map", {}))
            if scope_key_main == "mfk" and ip_counts_precomputed:
                ip_counts_precomputed = {(k[0], normalize_rname_for_scope("mfk", k[1])): v for k, v in ip_counts_precomputed.items()}
            if scope_key_main == "hermes" and ip_counts_precomputed:
                ip_counts_precomputed = {(k[0], normalize_rname_for_scope("hermes", k[1])): v for k, v in ip_counts_precomputed.items()}
            cache_loaded = True
            print("✅ Loaded input counts cache", flush=True)
        else:
            print("ℹ️ Cache version mismatch, rebuilding input counts", flush=True)
    except Exception as e:
        print(f"⚠️ Failed to load input counts cache. {str(e)}", flush=True)

if is_excel_display(IP_FILE):
    excel_path = excel_compatible_path(IP_FILE)
    if excel_path is not None:
        t0 = time.perf_counter()
        use_pandas_excel = os.getenv("PDP_EXCEL_PANDAS", "0").strip() == "1"
        scopes_for_count = None if (want_full_cache and not cache_loaded) else scopes_in_output
        collect_input_map = scope_key_main in {"mfk", "hermes"}
        if use_pandas_excel and not cache_loaded:
            ip_counts_precomputed, total_rows, kept_rows, ip_cols, ip_input_map = compute_ip_counts_pandas_excel(
                excel_path, scopes_for_count, collect_input_map=collect_input_map
            )
            if ip_counts_precomputed is None:
                print("⚠️ Pandas Excel load failed to find scope/rname columns. Falling back to streaming.", flush=True)
                ip_counts_precomputed, total_rows, kept_rows, ip_cols, ip_input_map = compute_ip_counts_excel(
                    excel_path, scopes_for_count, collect_input_map=collect_input_map
                )
        elif not cache_loaded:
            ip_counts_precomputed, total_rows, kept_rows, ip_cols, ip_input_map = compute_ip_counts_excel(
                excel_path, scopes_for_count, collect_input_map=collect_input_map
            )
    else:
        ip_counts_precomputed, total_rows, kept_rows, ip_cols = None, None, None, None
    if ip_counts_precomputed is None:
        print("⚠️ Excel streaming failed to find scope/rname columns. Falling back to full load.", flush=True)
    elif not cache_loaded:
        print(f"✅ Streamed IP in {time.perf_counter() - t0:.2f}s | rows={total_rows} | kept={kept_rows} | cols={ip_cols}", flush=True)

if ip_counts_precomputed is None:
    t0 = time.perf_counter()
    ip_df = read_file(IP_FILE)
    print(f"✅ Loaded IP in {time.perf_counter() - t0:.2f}s | rows={len(ip_df)}", flush=True)

    # Normalize column names
    ip_df.columns = [c.strip().lower() for c in ip_df.columns]

    # Normalize common alternate column names
    if "scope" not in ip_df.columns and "scope_name" in ip_df.columns:
        print("ℹ️ Renaming scope_name -> scope in INPUT file")
        ip_df = ip_df.rename(columns={"scope_name": "scope"})

    # Filter input by scopes found in output
    if "scope" in ip_df.columns:
        ip_df = ip_df[ip_df["scope"].isin(scopes_in_output)]
    print(f"✅ IP rows after scope filter: {len(ip_df)}", flush=True)

print(f"✅ Files loaded | OP Rows: {len(df)} | IP Rows: {len(ip_df) if ip_df is not None else 'streamed'} | Master Rows: {len(master_df)}")

# Write cache if requested and computed
if cache_write and ip_counts_precomputed is not None and not cache_loaded:
    try:
        payload = {
            "version": CACHE_VERSION,
            "counts": encode_counts(ip_counts_precomputed),
            "input_map": encode_input_map(ip_input_map)
        }
        with open(cache_write, "w", encoding="utf-8") as f:
            json.dump(payload, f)
        print("✅ Input counts cache written", flush=True)
    except Exception as e:
        print(f"⚠️ Failed to write input counts cache. {str(e)}", flush=True)

# ================= MASTER MAP (scope → rname/country) =================
valid_scope_rname = set()
valid_scope_country = set()

for _, r in master_df.iterrows():
    scope = str(r.get("scope", "")).strip()
    rname = str(r.get("rname", "")).strip()
    country = str(r.get("country", "")).strip()
    if scope and rname:
        valid_scope_rname.add((scope, rname))
    if scope and country:
        valid_scope_country.add((scope, country))

# ================= UNIQUE KEY =================
base_id = col_series(df, "base_id").astype(str).str.strip()
rname_key = col_series(df, "rname").astype(str).str.strip()
country_key = col_series(df, "country").astype(str).str.strip()

df["unique_key"] = rname_key + base_id

# Hermes special case: rname + base_id + country
mask_hermes = rname_key.str.upper().str.contains("HERMES_EUROPE", na=False)
df.loc[mask_hermes, "unique_key"] = rname_key[mask_hermes] + base_id[mask_hermes] + country_key[mask_hermes]

# ================= RNAME / COUNTRY CHECK =================
scope_s = col_series(df, "scope").astype(str).str.strip()
rname_s = col_series(df, "rname").astype(str).str.strip()
country_s = col_series(df, "country").astype(str).str.strip()

rname_pairs = pd.Series(list(zip(scope_s, rname_s)))
country_pairs = pd.Series(list(zip(scope_s, country_s)))

df["rname_check"] = np.where(rname_pairs.isin(valid_scope_rname), "PASS", "FAIL")
df["country_check"] = np.where(country_pairs.isin(valid_scope_country), "PASS", "FAIL")

# ================= ROW COUNT CHECK (per scope + rname) =================
row_key = list(zip(scope_s, rname_s.apply(lambda x: normalize_rname_for_scope(scope_key_main, x))))
op_counts = pd.Series(row_key).value_counts().to_dict()

if ip_counts_precomputed is not None:
    ip_counts = ip_counts_precomputed
else:
    ip_scope_col = "scope" if (ip_df is not None and "scope" in ip_df.columns) else None
    ip_rname_col = None
    if ip_df is not None and "rname" in ip_df.columns:
        ip_rname_col = "rname"
    elif ip_df is not None and "domain_input" in ip_df.columns:
        ip_rname_col = "domain_input"

    if ip_scope_col and ip_rname_col:
        ip_scope_vals = ip_df[ip_scope_col].astype(str).str.strip()
        ip_rname_vals = ip_df[ip_rname_col].astype(str).str.strip()
        ip_key = list(zip(
            ip_scope_vals,
            ip_rname_vals.apply(lambda x: normalize_rname_for_scope(scope_key_main, x))
        ))
        ip_counts = pd.Series(ip_key).value_counts().to_dict()
    else:
        ip_counts = {}

# If input map was not computed, try to derive from ip_df (non-excel path)
if not ip_input_map and ip_df is not None and scope_key_main in {"mfk", "hermes"}:
    if {"rname","base_id","country"}.issubset(ip_df.columns):
        rname_s = ip_df["rname"].astype(str).str.strip()
        base_s = ip_df["base_id"].astype(str).str.strip()
        country_s = ip_df["country"].astype(str).str.strip()
        pname_s = ip_df["pname"].astype(str).str.strip() if "pname" in ip_df.columns else pd.Series([""] * len(ip_df))
        skuvar_s = ip_df["skuvarient"].astype(str).str.strip() if "skuvarient" in ip_df.columns else pd.Series([""] * len(ip_df))
        for rn, b, c, pn, sv in zip(rname_s, base_s, country_s, pname_s, skuvar_s):
            key = f"{rn}{b}{c}"
            if key not in ip_input_map:
                ip_input_map[key] = {"pname": pn, "skuvarient": sv}

expected_counts = [ip_counts.get(k, 0) for k in row_key]
actual_counts = [op_counts.get(k, 0) for k in row_key]

df["row_count_input"] = expected_counts
df["row_count_output"] = actual_counts
df["row_count_check"] = np.where(df["row_count_input"] == df["row_count_output"], "PASS", "FAIL")

# ================= DATE CHECK =================
today = datetime.today().strftime("%Y-%m-%d")
if "date" in df.columns:
    df["date_check"] = df["date"].apply(lambda x: "PASS" if str(x).startswith(today) else "FAIL")

# ================= PRICE / STATUS CHECK =================
reg = col_series(df, "regularprice").astype(str).str.strip()
final = col_series(df, "finalprice").astype(str).str.strip()
markdown = col_series(df, "markdown_price").astype(str).str.strip()
item_status = col_series(df, "item_status").astype(str).str.strip()

reg_na = reg.apply(is_na_text) | reg.apply(is_not_available)
final_na = final.apply(is_na_text) | final.apply(is_not_available)

df["regular_final_match"] = np.where(reg == final, "TRUE", "FALSE")
df["price_rule_check"] = "PASS"

# Case 1: regular == final and both are NA/Not available
mask_na_eq = (reg == final) & reg_na & final_na
df.loc[mask_na_eq & (item_status != final), "price_rule_check"] = "FAIL"
df.loc[mask_na_eq & (markdown != final), "price_rule_check"] = "FAIL"

# Case 2: regular == final and NOT NA
mask_eq = (reg == final) & ~mask_na_eq
df.loc[mask_eq & (item_status != "R"), "price_rule_check"] = "FAIL"
df.loc[mask_eq & (~markdown.apply(is_na_text)), "price_rule_check"] = "FAIL"

# Case 3: regular != final
mask_neq = reg != final
df.loc[mask_neq & (item_status != "M"), "price_rule_check"] = "FAIL"
df.loc[mask_neq & (markdown.apply(is_na_text)), "price_rule_check"] = "FAIL"

# Additional rule: if item_status is n/a, then regularprice, finalprice, markdown_price must all be n/a to pass
mask_item_na = item_status.apply(is_na_text)
mask_prices_all_na = reg.apply(is_na_text) & final.apply(is_na_text) & markdown.apply(is_na_text)
df.loc[mask_item_na & ~mask_prices_all_na, "price_rule_check"] = "FAIL"

# Additional rule: if item_status is M, markdown_price cannot be n/a
mask_item_m = item_status.astype(str).str.strip() == "M"
df.loc[mask_item_m & markdown.apply(is_na_text), "price_rule_check"] = "FAIL"

# Additional rule: if item_status is R, markdown_price must be n/a
mask_item_r = item_status.astype(str).str.strip() == "R"
df.loc[mask_item_r & ~markdown.apply(is_na_text), "price_rule_check"] = "FAIL"

# Additional rule: regularprice must be >= finalprice unless both are n/a or Not Available
reg_num = pd.to_numeric(reg.where(~reg.apply(is_na_text) & ~reg.apply(is_not_available), None), errors="coerce")
final_num = pd.to_numeric(final.where(~final.apply(is_na_text) & ~final.apply(is_not_available), None), errors="coerce")
both_na = (reg.apply(is_na_text) | reg.apply(is_not_available)) & (final.apply(is_na_text) | final.apply(is_not_available))
df.loc[~both_na & (reg_num < final_num), "price_rule_check"] = "FAIL"

# Additional rule: if stock_status is In Stock, regular/final price cannot be n/a; Out of Stock can be n/a
stock_status_s = col_series(df, "stock_status").astype(str).str.strip()
mask_instock = stock_status_s == "In Stock"
mask_reg_na = reg.apply(is_na_text)
mask_final_na = final.apply(is_na_text)
df.loc[mask_instock & (mask_reg_na | mask_final_na), "price_rule_check"] = "FAIL"

# ================= STOCK / AVAILABILITY CHECK =================
stock_status = col_series(df, "stock_status").astype(str).str.strip()
availability = col_series(df, "availability").astype(str).str.strip()

df["availability_check"] = "PASS"

mask_in = stock_status == "In Stock"
mask_out = stock_status == "Out of Stock"

df.loc[mask_in & (availability != "Yes"), "availability_check"] = "FAIL"
df.loc[mask_out & (availability != "No"), "availability_check"] = "FAIL"

# If In Stock / Out of Stock, no "Not available" in key columns
def is_not_available_series(s: pd.Series) -> pd.Series:
    if isinstance(s, pd.DataFrame):
        s = s.iloc[:, 0]
    return s.astype(str).str.strip().str.lower() == "not available"

def is_na_or_not_available_series(s: pd.Series) -> pd.Series:
    return series_is_na(s) | is_not_available_series(s)

na_cols = []
for c in ["regularprice", "finalprice", "markdown_price", "item_status", "rating", "review", "availability"]:
    if c in df.columns:
        na_cols.append(c)

if na_cols:
    na_mask = np.zeros(len(df), dtype=bool)
    for c in na_cols:
        na_mask |= is_not_available_series(df[c])
    df["stock_na_check"] = np.where((mask_in | mask_out) & na_mask, "FAIL", "PASS")
else:
    df["stock_na_check"] = "PASS"

# ================= PID NOT NA CHECK =================
df["pid_check"] = "PASS"
pid_col = None
for c in df.columns:
    if c.lower() == "pid":
        pid_col = c
        break
if pid_col:
    pid_vals = col_series(df, pid_col).astype(str).str.strip()
    # n/a is NOT allowed (treat as missing)
    pid_bad = pid_vals.str.lower().isin({"", "n/a", "na", "null", "nan"})
    df.loc[pid_bad, "pid_check"] = "FAIL"

# ================= RNAME NOT NA CHECK (merge into rname_check) =================
rname_col = None
for c in df.columns:
    if c.lower() == "rname":
        rname_col = c
        break
if rname_col:
    rname_vals = col_series(df, rname_col).astype(str).str.strip()
    rname_bad = rname_vals.str.lower().isin({"", "n/a", "na", "null", "nan"})
    df.loc[rname_bad, "rname_check"] = "FAIL"

# ================= REQUIRED NON-NA FIELDS CHECK =================
REQUIRED_NON_NA_COLS = [
    "country","pname","brand","availability","rating","review","productimage",
    "date","scope","stock_status","multiple_url","large_image","small_image",
    "spid","item_status"
]
df["required_non_na_check"] = "PASS"
df["required_non_na_detail"] = ""

# benefit alias for country already mapped to country; also handle case-insensitive lookups
for req in REQUIRED_NON_NA_COLS:
    col_match = None
    for c in df.columns:
        if c.lower() == req:
            col_match = c
            break
    if not col_match:
        continue
    vals = col_series(df, col_match).astype(str).str.strip()
    # only treat explicit n/a values as invalid (do NOT treat blanks as n/a)
    bad = vals.str.lower().isin({"n/a","na","null","nan"})
    # special case for item_status: allow n/a only if all prices are n/a
    if col_match.lower() == "item_status":
        prices_all_na = reg.apply(is_na_text) & final.apply(is_na_text) & markdown.apply(is_na_text)
        bad = bad & ~prices_all_na
    if bad.any():
        df.loc[bad, "required_non_na_check"] = "FAIL"
        df.loc[bad, "required_non_na_detail"] = df.loc[bad, "required_non_na_detail"] + f"{col_match} cant have n/a values | "

# Hermes only: currency cannot be n/a
scope_key_for_na = scope_key_from_value(col_series(df, "scope", "").astype(str).str.strip().iloc[0] if len(df) else "")
if scope_key_for_na == "hermes":
    currency_col = None
    for c in df.columns:
        if c.lower() == "currency":
            currency_col = c
            break
    if currency_col:
        cur_vals = col_series(df, currency_col).astype(str).str.strip()
        cur_bad = cur_vals.str.lower().isin({"n/a","na","null","nan"})
        if cur_bad.any():
            df.loc[cur_bad, "required_non_na_check"] = "FAIL"
            df.loc[cur_bad, "required_non_na_detail"] = df.loc[cur_bad, "required_non_na_detail"] + f"{currency_col} cant have n/a values | "

# Hermes only: UPC must not be in scientific notation
if scope_key_for_na == "hermes":
    df["upc_check"] = "PASS"
    upc_col = None
    for c in df.columns:
        if c.lower() == "upc":
            upc_col = c
            break
    if upc_col:
        upc_vals = col_series(df, upc_col).astype(str).str.strip()
        sci_mask = upc_vals.str.contains(r"[eE]\+?\d+", regex=True, na=False)
        if sci_mask.any():
            df.loc[sci_mask, "upc_check"] = "FAIL"
            df.loc[sci_mask, "required_non_na_check"] = "FAIL"
            df.loc[sci_mask, "required_non_na_detail"] = df.loc[sci_mask, "required_non_na_detail"] + f"{upc_col} in scientific notation | "

# ================= NOT AVAILABLE VALUE CHECK =================
NOT_AVAILABLE_REQUIRED_BY_SCOPE = {
    "hermes": ["top_category","category","sub_category","RName","country","SKUVARIENT","PName","Brand","Date","scope","base_id"],
    "coty": ["top_category","category","sub_category","RName","country","SKUVARIENT","Date","scope","base_id","url"],
    "mfk": ["top_category","category","sub_category","RName","country","SKUVARIENT","PName","Brand","Date","scope","base_id"],
    "pcd": ["top_category","category","sub_category","RName","country","SKUVARIENT","PName","Brand","Date","scope","base_id","url"],
}
ALLOWED_NA_REQUIRED_BY_SCOPE = {
    "pcd": {"url"},
}

scope_name = col_series(df, "scope", "").astype(str).str.strip().iloc[0] if len(df) else ""
scope_key = scope_key_from_value(scope_name)
required_na_cols = NOT_AVAILABLE_REQUIRED_BY_SCOPE.get(scope_key)
df["not_available_values_check"] = "PASS"
df["not_available_values_detail"] = ""

if required_na_cols:
    na_mask = stock_status.str.strip().str.lower() == "not available"
    if na_mask.any():
        col_map = {c.lower(): c for c in base_columns}
        required_na_lower = [c.lower() for c in required_na_cols]
        allowed_na = {c.lower() for c in ALLOWED_NA_REQUIRED_BY_SCOPE.get(scope_key, set())}

        required_actual = []
        missing_required = []
        for req_l in required_na_lower:
            actual = col_map.get(req_l)
            if actual:
                required_actual.append(actual)
            else:
                missing_required.append(req_l)

        fail_mask = pd.Series(False, index=df.index)
        detail = pd.Series([""] * len(df), index=df.index, dtype=object)
        if missing_required:
            fail_mask |= na_mask

        # Required columns should NOT be "Not available" (NA is allowed)
        for col in required_actual:
            if col.lower() in allowed_na:
                continue
            mask_bad = na_mask & is_not_available_series(df[col])
            fail_mask |= mask_bad
            if mask_bad.any():
                detail.loc[mask_bad] += f"required {col} is Not available | "

        # Non-required columns should be not available
        other_cols = [c for c in base_columns if c.lower() not in set(required_na_lower)]
        for col in other_cols:
            mask_bad = na_mask & ~is_na_or_not_available_series(df[col])
            fail_mask |= mask_bad
            if mask_bad.any():
                detail.loc[mask_bad] += f"non-required {col} has value | "

        df.loc[fail_mask, "not_available_values_check"] = "FAIL"
        df.loc[fail_mask, "not_available_values_detail"] = detail.str.rstrip(" | ")

# ================= RATING CHECK =================
rating = col_series(df, "rating").astype(str).str.strip()
review = col_series(df, "review").astype(str).str.strip()

df["rating_check"] = "PASS"

rating_is_na = rating.apply(is_na_text) | rating.apply(is_not_available)
review_is_na = review.apply(is_na_text) | review.apply(is_not_available)
# normalize decimal comma and remove stray spaces
rating_norm = rating.str.replace(",", ".", regex=False).str.strip()
rating_norm = rating_norm.replace({"": np.nan})
rating_numeric_ok = rating_norm.str.match(r"^[0-9]+(\.[0-9]+)?$", na=False)

rating_val = pd.to_numeric(rating_norm.where(rating_numeric_ok, np.nan), errors="coerce")
rating_in_range = (rating_val >= 0) & (rating_val <= 5)

review_val = pd.to_numeric(review, errors="coerce").fillna(0)

# If stock status is Not available, allow text rating
mask_stock_na = stock_status == "Not available"

# Exceptional case: rating and review both Not available -> PASS
mask_rating_review_na = rating.apply(is_not_available) & review.apply(is_not_available)
df.loc[mask_rating_review_na, "rating_check"] = "PASS"

# For non-Not-available stock, rating must be numeric 0-5 and no spaces
df.loc[~mask_stock_na & (~rating_numeric_ok | ~rating_in_range) & ~mask_rating_review_na, "rating_check"] = "FAIL"

# If review > 0, rating must be present
df.loc[(review_val > 0) & rating_is_na & ~mask_rating_review_na, "rating_check"] = "FAIL"

# For MFK and Hermes: In Stock / Out of Stock must not have Not available in rating or review
scope_key_rr = scope_key_from_value(col_series(df, "scope", "").astype(str).str.strip().iloc[0] if len(df) else "")
if scope_key_rr in {"mfk", "hermes"}:
    df["rating_review_stock_check"] = "PASS"
    mask_stock_in_out = stock_status.isin(["In Stock", "Out of Stock"])
    mask_rating_na = rating.apply(is_not_available)
    mask_review_na = review.apply(is_not_available)
    df.loc[mask_stock_in_out & (mask_rating_na | mask_review_na), "rating_review_stock_check"] = "FAIL"

# For MFK and Hermes: Not available SKUs must have PName and SKUVARIENT from input (not Not available)
scope_key_rr = scope_key_from_value(col_series(df, "scope", "").astype(str).str.strip().iloc[0] if len(df) else "")
if scope_key_rr in {"mfk", "hermes"} and ip_input_map:
    df["na_sku_input_check"] = "PASS"
    mask_na_sku = stock_status.str.strip().str.lower() == "not available"
    if mask_na_sku.any():
        pname_col = None
        skuvar_col = None
        for c in df.columns:
            if c.lower() == "pname":
                pname_col = c
            elif c.lower() == "skuvarient":
                skuvar_col = c
        if pname_col is None or skuvar_col is None:
            df.loc[mask_na_sku, "na_sku_input_check"] = "FAIL"
        else:
            out_rname = col_series(df, "rname").astype(str).str.strip()
            out_base = col_series(df, "base_id").astype(str).str.strip()
            out_country = col_series(df, "country").astype(str).str.strip()
            out_pname = col_series(df, pname_col).astype(str).str.strip()
            out_skuvar = col_series(df, skuvar_col).astype(str).str.strip()
            for i in df[mask_na_sku].index:
                key = f"{out_rname[i]}{out_base[i]}{out_country[i]}"
                ref = ip_input_map.get(key)
                if not ref:
                    df.at[i, "na_sku_input_check"] = "FAIL"
                    continue
                if is_na_text(out_pname[i]) or is_not_available(out_pname[i]) or is_na_text(out_skuvar[i]) or is_not_available(out_skuvar[i]):
                    df.at[i, "na_sku_input_check"] = "FAIL"
                    continue
                if ref.get("pname", "") and out_pname[i] != ref.get("pname", ""):
                    df.at[i, "na_sku_input_check"] = "FAIL"
                    continue
                if ref.get("skuvarient", "") and out_skuvar[i] != ref.get("skuvarient", ""):
                    df.at[i, "na_sku_input_check"] = "FAIL"

# For MFK and Hermes: Not available SKU must have url = "Not Available" (not n/a)
if scope_key_rr in {"mfk", "hermes"}:
    df["na_url_check"] = "PASS"
    mask_na_sku = stock_status.str.strip().str.lower() == "not available"
    if mask_na_sku.any():
        url_col = None
        for c in df.columns:
            if c.lower() == "url":
                url_col = c
                break
        if url_col is None:
            df.loc[mask_na_sku, "na_url_check"] = "FAIL"
        else:
            url_vals = col_series(df, url_col).astype(str).str.strip()
            bad = mask_na_sku & ~url_vals.str.lower().eq("not available")
            df.loc[bad, "na_url_check"] = "FAIL"

# Rating normalization for MFK and Hermes (check only, no mutation)
scope_key_rating = scope_key_from_value(col_series(df, "scope", "").astype(str).str.strip().iloc[0] if len(df) else "")
if scope_key_rating in {"mfk", "hermes"}:
    df["rating_normalization_check"] = "PASS"
    mask_rating_valid = rating_numeric_ok & rating_in_range
    mask_rating_zero = mask_rating_valid & (rating_val == 0)
    mask_rating_int = mask_rating_valid & (rating_val % 1 == 0) & ~mask_rating_zero
    # Fail if rating is integer > 0 but not written as X.0, or if rating is invalid
    normalized_int_ok = rating_norm.str.match(r"^[0-9]+\.0+$", na=False)
    normalized_zero_ok = rating_norm.eq("0")
    fail_norm = (mask_rating_int & ~normalized_int_ok) | (mask_rating_zero & ~normalized_zero_ok)
    fail_norm |= (~mask_rating_valid & ~rating_is_na & ~mask_rating_review_na)
    df.loc[fail_norm, "rating_normalization_check"] = "FAIL"

# ================= KEYWORDS BLANK CHECK (scope-specific) =================
KEYWORDS_BLANK_SCOPES = {"mfk", "hermes"}
scope_key = scope_key_from_value(col_series(df, "scope", "").astype(str).str.strip().iloc[0] if len(df) else "")
if scope_key in KEYWORDS_BLANK_SCOPES:
    df["keywords_blank_check"] = "PASS"
    # find keywords column case-insensitively
    keywords_col = None
    for c in df.columns:
        if c.lower() == "keywords":
            keywords_col = c
            break
    if keywords_col is None:
        df["keywords_blank_check"] = "FAIL"
    else:
        kw_vals = col_series(df, keywords_col).astype(str).str.strip()
        # must be truly blank; n/a or not available should FAIL
        non_blank = kw_vals != ""
        df.loc[non_blank, "keywords_blank_check"] = "FAIL"

# ================= FAILURE REASON =================
REQUIRED_COLUMNS_BY_SCOPE = {
    # MFK and PCD scopes
    "MFK": [
        "top_category","category","sub_category","pid","key","RName","country","MPN","SKUVARIENT","SKUL","PName",
        "Division","Category","Department","Class","SubClass","Brand","RegularPrice","FinalPrice","PriceRange",
        "Availability","Rating","Review","Promotion","Position","Channel","currency","ProductImage","KeyWords",
        "day","month","year","Date","scope","Product_Description","base_id","region_input","currency_input",
        "stock_status","item_status","markdown_price","url","category_path","add_to_cart","multiple_url",
        "large_image","small_image","normalized_Brand","spid"
    ],
    "PCD": [
        "top_category","category","sub_category","pid","key","RName","country","MPN","SKUVARIENT","SKUL","PName",
        "Division","Category","Department","Class","SubClass","Brand","RegularPrice","FinalPrice","PriceRange",
        "Availability","Rating","Review","Promotion","Position","Channel","currency","ProductImage","KeyWords",
        "day","month","year","Date","scope","Product_Description","base_id","region_input","currency_input",
        "stock_status","item_status","markdown_price","url","category_path","add_to_cart","multiple_url",
        "large_image","small_image","normalized_Brand","spid"
    ],
    "Benefit": [
        "site_name","category_path","category_path_url","division","category","department","class","subclass",
        "product_id","product_name","product_description","product_dimensions","product_weight","product_material",
        "url","product_image","regular_price","regular_price_range","shipping_price","markdown_price",
        "disounted_price","final_price","item_status","item_level_status","features","color","price_by_size",
        "additional_information","promo_message","price_promo","promo_description","online_exclusive",
        "extraction_date","image_url_large","image_url_small","base_unique_identifier","mpn","page_title",
        "pack_quantity","brand","shipping_text","reviews_count","average_rating","upc","technical_details",
        "meta_keywords","stock_availability","page_snapshot","item_condition","no_of_watchers","stock_status",
        "shipping_areas","shipping_weight","meta_description","canonical_url","meta_title","videos","add_on_item",
        "saleable_quantity","ship_surcharge","in_cart_price","add_to_cart_price","sku_variant","sku_id",
        "variation_url","availability","Frequency","UOM","region_site","region_input","base_id","currency_input",
        "scope","top_category","input_category","sub_category","images_chunk","Retailer","Project","Batch",
        "SubBatch","input_url","spid","normalizedpname","top_sku","brand_Input","brand_group","sku_type",
        "buffer_column_1","buffer_column_2","buffer_column_3","buffer_column_4","buffer_column_5"
    ],
    "COTY": [
        "top_category","category","sub_category","pid","key","RName","country","MPN","SKUVARIENT","SKUL","PName",
        "Division","Category","Department","Class","SubClass","Brand","RegularPrice","FinalPrice","PriceRange",
        "Availability","Rating","Review","Promotion","Position","Channel","currency","ProductImage","KeyWords",
        "day","month","year","Date","scope","Product_Description","base_id","region_input","currency_input",
        "stock_status","item_status","markdown_price","url","category_path","add_to_cart","multiple_url",
        "large_image","small_image","normalized_Brand","spid","video_from_carousel","video_from_body",
        "pdp+_content","features","specifications","seller_name","asin","upc"
    ],
    "Hermes": [
        "top_category","category","sub_category","pid","key","RName","country","MPN","SKUVARIENT","SKUL","PName",
        "Division","Category","Department","Class","SubClass","Brand","RegularPrice","FinalPrice","PriceRange",
        "Availability","Rating","Review","Promotion","Position","Channel","currency","ProductImage","KeyWords",
        "day","month","year","Date","scope","Product_Description","base_id","region_input","currency_input",
        "stock_status","item_status","markdown_price","url","category_path","add_to_cart","multiple_url",
        "large_image","small_image","normalized_Brand","spid","video_from_carousel","video_from_body",
        "pdp+_content","features","specifications","seller_name","asin","upc"
    ]
}

scope_values = set(df.get("scope", pd.Series([""])).astype(str).str.strip())
required_columns = None
for s in scope_values:
    scope_key = scope_key_from_value(s)
    for k, cols in REQUIRED_COLUMNS_BY_SCOPE.items():
        if k.lower() == scope_key:
            required_columns = cols
            break
    if required_columns:
        break

if required_columns:
    op_cols = get_file_columns(OP_FILE, display_override=display_name(OP_FILE))
    current_cols = op_cols if op_cols else list(df.columns)
    # case-sensitive checks against original headers
    missing_required = [req for req in required_columns if req not in current_cols]
    extra_current = [cur for cur in current_cols if cur not in set(required_columns)]
    # sequence check: compare the sequence of required columns as they appear in current columns (case-sensitive)
    current_required_sequence = [c for c in current_cols if c in set(required_columns)]
    sequence_ok = current_required_sequence == required_columns
    df["required_column_check"] = "PASS" if (len(missing_required) == 0 and sequence_ok) else "FAIL"
    df["missing_columns_check"] = "PASS" if len(missing_required) == 0 else "FAIL"
    df["extra_columns_check"] = "PASS" if len(extra_current) == 0 else "FAIL"
else:
    df["required_column_check"] = "PASS"
    df["missing_columns_check"] = "PASS"
    df["extra_columns_check"] = "PASS"

FAILURE_MESSAGE_MAP = {
    "rname_check": "Retailer name (rname) is not valid for this scope (check master mapping).",
    "country_check": "Country is not valid for this scope (check master mapping).",
    "row_count_check": "Input row count does not match output row count for this scope + rname.",
    "date_check": "Date is not today (expected current date).",
    "price_rule_check": "Price/Status rule failed: regular vs final price must match item_status and markdown_price.",
    "availability_check": "Availability must match stock_status: In Stock -> Yes, Out of Stock -> No.",
    "stock_na_check": "For In Stock/Out of Stock, these columns cannot be 'Not available': regularprice/finalprice/markdown_price/item_status/rating/review/availability.",
    "rating_check": "Rating is invalid: must be numeric 0–5 (or allowed NA for Not available stock).",
    "keywords_blank_check": "Keywords must be blank for this scope (MFK/Hermes).",
    "required_column_check": "Required columns missing or wrong order (scope-specific template mismatch).",
    "missing_columns_check": "One or more required columns are missing for this scope.",
    "extra_columns_check": "One or more extra columns exist that are not in the required template.",
    "not_available_values_check": "Not available rule failed: when stock_status = 'Not available', only required columns may have values; all other columns must be NA/Not available.",
    "pid_check": "PID is missing or n/a.",
    "rname_check": "Retailer name (rname) is not valid for this scope or is missing/n/a.",
    "required_non_na_check": "One or more required fields contain n/a values.",
    "rating_normalization_check": "Rating format rule failed (MFK/Hermes): integers must be X.0 and zero must be '0'.",
    "rating_review_stock_check": "For In Stock/Out of Stock (MFK/Hermes), rating and review cannot be 'Not available'.",
    "na_sku_input_check": "Not available SKU must match input PName and SKUVARIENT (MFK/Hermes, key=rname+base_id+country).",
    "na_url_check": "For Not Available SKU (MFK/Hermes), url must be 'Not Available' (not n/a).",
    "upc_check": "UPC is in scientific notation (not allowed).",
}

check_cols = [c for c in FAILURE_MESSAGE_MAP.keys() if c in df.columns]

df["failure_reason"] = ""
for col in check_cols:
    df.loc[df[col].eq("FAIL"), "failure_reason"] += FAILURE_MESSAGE_MAP[col] + " | "

df["failure_reason"] = df["failure_reason"].str.rstrip(" | ")
df["overall_status"] = np.where(df["failure_reason"] == "", "PASS", "FAIL")

# Append detailed non-NA failure messages if present
if "required_non_na_detail" in df.columns:
    detail = df["required_non_na_detail"].str.rstrip(" | ")
    df.loc[detail != "", "failure_reason"] = df.loc[detail != "", "failure_reason"] + " | " + detail
    df["failure_reason"] = df["failure_reason"].str.strip(" | ")
    df["overall_status"] = np.where(df["failure_reason"] == "", "PASS", "FAIL")

# ================= OUTPUT =================
with pd.ExcelWriter(OUTPUT_FILE, engine="xlsxwriter", engine_kwargs={"options": {"strings_to_urls": False}}) as writer:
    df.to_excel(writer, sheet_name="PDP_Data", index=False)
    if required_columns:
        op_cols = get_file_columns(OP_FILE, display_override=display_name(OP_FILE))
        current_cols = op_cols if op_cols else list(df.columns)
        req_df = pd.DataFrame([{
            "scope": next(iter(scope_values)) if scope_values else "",
            "required_columns": ", ".join(required_columns),
            "current_columns": ", ".join(current_cols),
            "overall_status": "FAIL" if df["required_column_check"].iloc[0] == "FAIL" else "PASS",
            "failure_message": "all required columns not present or order mismatch" if df["required_column_check"].iloc[0] == "FAIL" else ""
        }])
        req_df.to_excel(writer, sheet_name="Required_Columns_Check", index=False)
        detail_df = pd.DataFrame([{
            "scope": next(iter(scope_values)) if scope_values else "",
            "missing_columns": ", ".join(missing_required) if required_columns else "",
            "extra_columns": ", ".join(extra_current) if required_columns else "",
            "missing_columns_check": df["missing_columns_check"].iloc[0],
            "extra_columns_check": df["extra_columns_check"].iloc[0]
        }])
        detail_df.to_excel(writer, sheet_name="Missing_Extra_Columns", index=False)

print(f"✅ OUTPUT GENERATED: {OUTPUT_FILE}")
print("=== PDP CHECKLIST COMPLETED ===")
